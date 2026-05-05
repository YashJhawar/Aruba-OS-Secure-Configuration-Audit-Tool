import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
import re
import threading
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

# ─────────────────────────────────────────────
# AUDIT KNOWLEDGE BASE
# Source: CIS HPE Aruba Networking CX Switch Benchmark v1.0.1
# ─────────────────────────────────────────────
AUDIT_CHECKS = [

    # ══════════════════════════════════════════
    # SECTION 1 – MANAGEMENT
    # ══════════════════════════════════════════

    # ── 1.1 Securing Switch Management Access ─

    {
        "id": 1,
        "cis_id": "1.1.3",
        "obs_title": "Hardening Password Rules (Automated)",
        "observation": "It was observed that password complexity rules are not configured on the device.",
        "severity": "High",
        "description": (
            "It is critical that customers set good, strong passwords for their local user accounts. "
            "The password complexity feature can aid customers in ensuring that users are following "
            "company password policies. Password complexity allows administrators to define a set of "
            "password rules which can then align with their company's password policy. The password "
            "complexity allows for minimum password length, minimum lowercase/uppercase/special/"
            "numeric characters, position changes between passwords, and can protect against password "
            "reuse. CIS recommends 14 characters or more in length with 1 Uppercase, 1 Lowercase, "
            "1 Number and 1 Special Character."
        ),
        "impact": (
            "Use of weak or easily guessed passwords can result in unintended access of the device "
            "and could lead to someone taking over the network. There are many known attack vectors "
            "against weak passwords including brute force, dictionary, and hybrid attacks."
        ),
        "recommendation": (
            "It is recommended to configure password complexity rules.\n\n"
            "Set a password complexity policy with a minimum of <X> characters long, includes at "
            "least <Y> lowercase/uppercase/special/numeric character, and doesn't allow users to "
            "reuse the past <Z> passwords:\n\n"
            "switch(config)# password complexity\n"
            "switch(config-pwd-cplx)# minimum-length <X>\n"
            "switch(config-pwd-cplx)# lowercase-count <Y>\n"
            "switch(config-pwd-cplx)# uppercase-count <Y>\n"
            "switch(config-pwd-cplx)# special-char-count <Y>\n"
            "switch(config-pwd-cplx)# numeric-count <Y>\n"
            "switch(config-pwd-cplx)# history-count <Z>\n"
            "switch(config-pwd-cplx)# enable"
        ),
        "expected_output": (
            "switch(config)# password complexity\n"
            "switch(config-pwd-cplx)# minimum-length <X>\n"
            "switch(config-pwd-cplx)# lowercase-count <Y>\n"
            "switch(config-pwd-cplx)# uppercase-count <Y>\n"
            "switch(config-pwd-cplx)# special-char-count <Y>\n"
            "switch(config-pwd-cplx)# numeric-count <Y>\n"
            "switch(config-pwd-cplx)# history-count <Z>\n"
            "switch(config-pwd-cplx)# enable"
        ),
        "check_fn": "check_password_complexity",
    },

    {
        "id": 2,
        "cis_id": "1.1.4",
        "obs_title": "Set an Export Password (Automated)",
        "observation": "It was observed that the device is using the default export password.",
        "severity": "Medium",
        "description": (
            "All AOS-CX switches ship with a default password used to export any customer secrets "
            "contained within the switch configuration. Display of secrets through 'show' commands "
            "or exporting the entire switch configuration will have its secrets encrypted with this "
            "password prior to being displayed. AOS-CX allows customers to set their own password "
            "to ensure that their switch secrets cannot be loaded onto another AOS-CX switch."
        ),
        "impact": (
            "Secrets are never exposed in plaintext, but loading of another switch's secrets onto "
            "a different switch could enable that device to impersonate another switch in the network. "
            "This may include, but is not limited to, MACsec peering with an MKA PSK or OSPFv2 "
            "shared secrets."
        ),
        "recommendation": (
            "It is recommended to set a custom export password.\n\n"
            "switch(config)# service export-password\n"
            "Enter password: ********\n"
            "Confirm password: ********"
        ),
        "expected_output": (
            "switch# show service export-password\n"
            "Export password: custom"
        ),
        "check_fn": "check_export_password",
    },

    {
        "id": 3,
        "cis_id": "1.1.8",
        "obs_title": "Session Management (Automated)",
        "observation": "It was observed that CLI session management limits are not configured on the device.",
        "severity": "Medium",
        "description": (
            "Session management enhances security by enforcing specific CLI user session requirements "
            "for console, SSH and Telnet connections. The following information is provided at the time "
            "of a successful login: number of failed login attempts since the most recent successful "
            "login, the date/time/location of the most recent previous successful login, and the count "
            "of successful logins within the past (configurable) time period. CIS recommends a minimum "
            "setting of: max-per-user 1, tracking-range 25, timeout 15."
        ),
        "impact": (
            "Limiting concurrent sessions and session timeouts is crucial for enhancing security and "
            "preventing unauthorized access to accounts. Failure to restrict the number of active "
            "sessions per user and setting appropriate timeout durations can expose the device to "
            "risks like session hijacking, shared account access, and compromised sessions."
        ),
        "recommendation": (
            "It is recommended to configure CLI session management.\n\n"
            "switch(config)# cli-session\n"
            "switch(config-cli-session)# max-per-user 1\n"
            "switch(config-cli-session)# timeout 15\n"
            "switch(config-cli-session)# tracking-range 25\n"
            "switch(config-cli-session)# exit"
        ),
        "expected_output": (
            "switch(config)# cli-session\n"
            "switch(config-cli-session)# max-per-user 1\n"
            "switch(config-cli-session)# timeout 15\n"
            "switch(config-cli-session)# tracking-range 25\n"
            "switch(config-cli-session)# exit"
        ),
        "check_fn": "check_session_management",
    },

    {
        "id": 4,
        "cis_id": "1.1.9",
        "obs_title": "Verifying Telnet Server is Disabled (Automated)",
        "observation": "It was observed that the Telnet server is enabled on the device.",
        "severity": "High",
        "description": (
            "AOS-CX system includes a Telnet server. Telnet is considered insecure primarily because "
            "it transmits all data, including usernames and passwords, in plaintext, leaving it "
            "vulnerable to eavesdropping, credential sniffing, Man-in-the-Middle (MitM) attacks, and "
            "session hijacking. It is recommended to use Secure Shell (SSH) instead of Telnet, which "
            "provides encryption for data transmission and strong authentication mechanisms, ensuring "
            "a much higher level of security for remote access."
        ),
        "impact": (
            "Verifying Telnet server is disabled will prevent users from accessing the console using "
            "an insecure protocol. Telnet server is included in AOS-CX for customer convenience but "
            "is not recommended for general use as it exposes all transmitted data in plaintext."
        ),
        "recommendation": (
            "It is recommended to disable the Telnet server.\n\n"
            "switch(config)# no telnet server vrf <vrf>\n"
            "All active TELNET sessions on the VRF will be terminated.\n"
            "Do you want to continue (y/n)? y"
        ),
        "expected_output": (
            "switch(config)# show telnet server\n"
            "TELNET server is not enabled on any VRFs."
        ),
        "check_fn": "check_telnet_disabled",
    },

    # ── 1.2 Securing SSH ───────────────────────

    {
        "id": 5,
        "cis_id": "1.2.1",
        "obs_title": "SSH Public Key Authentication (Automated)",
        "observation": "It was observed that SSH public key authentication is not configured for users.",
        "severity": "Medium",
        "description": (
            "SSH Public key authentication is enabled by default and takes precedence over "
            "password-based authentication. Passwords are vulnerable to attacks and human errors. "
            "Keys are more secure and efficient compared to passwords. Public-key authentication "
            "uses a cryptographic pair of keys (public and private) to verify user identity and "
            "secure communication. A user's private key is kept securely on the device used to "
            "connect to the switch. The user's public key is configured on the switch as an "
            "authorizing key for that user."
        ),
        "impact": (
            "Without SSH public key authentication, users rely solely on password-based "
            "authentication which is vulnerable to brute-force, credential stuffing, and "
            "man-in-the-middle attacks. Configuring public-key authentication significantly "
            "reduces these risks by leveraging cryptographic key pairs."
        ),
        "recommendation": (
            "It is recommended to configure SSH public key authentication for administrative users.\n\n"
            "switch(config)# user <username> authorized-key <user public-key>"
        ),
        "expected_output": (
            "switch# show ssh authentication-method\n"
            " SSH publickey authentication      : Enabled\n"
            " SSH password authentication       : Enabled"
        ),
        "check_fn": "check_ssh_pubkey",
    },

    {
        "id": 6,
        "cis_id": "1.2.2",
        "obs_title": "SSH Allow List (Automated)",
        "observation": "It was observed that SSH allow list is not enabled on the device.",
        "severity": "Medium",
        "description": (
            "SSH server allow-list allows the switch admin to limit in-bound SSH connections to "
            "be from specific IP addresses or networks. Configure a list of addresses that will be "
            "the only hosts allowed to connect to the SSH servers running on all VRFs of the switch. "
            "By default, the allow-list is disabled and any host is allowed to connect given the "
            "correct authentication criteria."
        ),
        "impact": (
            "When the allow-list is enabled, only the hosts that fall under one of the entries may "
            "connect with the correct authentication criteria; all other hosts will be denied to "
            "attempt authentication. Without this restriction, any IP address can attempt SSH "
            "authentication, increasing exposure to brute-force and credential-guessing attacks."
        ),
        "recommendation": (
            "It is recommended to enable SSH Allow List with permitted management station IPs.\n\n"
            "switch(config)# ssh server allow-list\n"
            "switch(config-ssh-al)# ip 1.1.1.1\n"
            "switch(config-ssh-al)# ip 2.2.2.0/24\n"
            "switch(config-ssh-al)# enable"
        ),
        "expected_output": (
            "switch# show ssh server allow-list\n"
            "SSH server allow-list:\n"
            "  Status: Enabled\n"
            "  Allowed host IPs:\n"
            "    1.1.1.1"
        ),
        "check_fn": "check_ssh_allowlist",
    },

    {
        "id": 7,
        "cis_id": "1.2.3",
        "obs_title": "SSH Server Port Customization (Automated)",
        "observation": "It was observed that SSH server is listening on the default TCP port 22.",
        "severity": "Low",
        "description": (
            "By default, SSH server listens on TCP port 22. This port will be used for all VRFs "
            "that have SSH server enabled. Optionally, AOS-CX switches provide the ability to "
            "modify the default SSH server port to add extra protection to the server. Configuring "
            "a non-default SSH port provides an additional layer of obscurity and reduces automated "
            "scanning activity targeting the default port."
        ),
        "impact": (
            "Configuring a SSH server listening port allows admins to provide SSH connectivity to "
            "the switch when network situations, such as business networks and WiFi hotspots, block "
            "IP port 22. It also reduces the volume of automated scanning and brute-force attempts "
            "targeting the well-known default port."
        ),
        "recommendation": (
            "It is recommended to configure a non-default SSH server port.\n\n"
            "switch(config)# ssh server port <non-default-port>"
        ),
        "expected_output": (
            "switch# show ssh server\n"
            "SSH server configuration on VRF default :\n"
            "    TCP Port          : <non-default-port>"
        ),
        "check_fn": "check_ssh_port",
    },

    {
        "id": 8,
        "cis_id": "1.2.5",
        "obs_title": "Two-factor authentication with the SSH server (Automated)",
        "observation": "It was observed that two-factor authentication is not enabled for the SSH server.",
        "severity": "Medium",
        "description": (
            "Two-factor authentication provides an additional level of security by requiring "
            "administrators to perform X.509 certificate-based authentication with the AOS-CX SSH "
            "server. The two factors are: possession of the certificate private key and knowledge "
            "of the password associated with the certificate private key. The switch must be "
            "configured with the corresponding CA certificate which issued the administrator's "
            "certificate."
        ),
        "impact": (
            "Any changes in administrator user passwords or the addition or removal of administrator "
            "accounts can become a maintenance challenge for customers. Use of two-factor "
            "authentication only requires the enablement of the feature and the addition of the "
            "corresponding CA certificate as a trust anchor, reducing operational overhead while "
            "significantly improving authentication security."
        ),
        "recommendation": (
            "It is recommended to enable two-factor authentication for the SSH server.\n\n"
            "switch(config)# ssh certificate-as-authorized-key\n"
            "switch(config)# ssh two-factor-authentication authorization local"
        ),
        "expected_output": (
            "switch(config)# show ssh authentication-method\n"
            " SSH publickey authentication      : Enabled\n"
            " SSH two factor authentication     : Enabled"
        ),
        "check_fn": "check_ssh_2fa",
    },

    # ── 1.3 Securing Time Synchronization ─────

    {
        "id": 9,
        "cis_id": "1.3.1",
        "obs_title": "NTP Authentication (Automated)",
        "observation": "It was observed that NTP authentication is not configured on the device.",
        "severity": "Medium",
        "description": (
            "NTP authentication support allows the NTP client to verify that servers are in fact "
            "known and trusted and not intruders intending to masquerade as a legitimate server. "
            "Many secure protocols and auditing functions rely on system times being synchronized "
            "with a reliable time source. Network Time Protocol (NTP) is used to synchronize clocks "
            "on networked devices, ensuring consistent timestamps for various applications."
        ),
        "impact": (
            "Without NTP authentication, attackers can disguise themselves as NTP servers and "
            "broadcast wrong time, making it difficult to correlate events during an incident. "
            "An attacker could potentially spoof NTP packets, leading to time-related attacks "
            "such as replay attacks or man-in-the-middle attacks."
        ),
        "recommendation": (
            "It is recommended to enable NTP authentication and set a key to authenticate NTP servers.\n\n"
            "switch(config)# ntp authentication\n"
            "switch(config)# ntp authentication-key 1 md5 <ntpauthkey>\n"
            "switch(config)# ntp server 10.10.10.10 prefer\n"
            "switch(config)# ntp vrf mgmt"
        ),
        "expected_output": (
            "switch(config)# ntp authentication\n"
            "switch(config)# ntp authentication-key 1 md5 <ntpauthkey>"
        ),
        "check_fn": "check_ntp_auth",
    },

    {
        "id": 10,
        "cis_id": "1.3.2",
        "obs_title": "Configuring Time Services (Automated)",
        "observation": "It was observed that NTP time services are not configured on the device.",
        "severity": "Low",
        "description": (
            "Network Time services are used to synchronize computer clocks across a network, "
            "ensuring accurate and consistent timekeeping for various applications and services. "
            "This synchronization is crucial for tasks like file system updates, network management, "
            "security, and distributed systems. AOS-CX supports both NTP client and NTP server "
            "functionality. Operational and security best practice is to configure more than "
            "1 NTP server."
        ),
        "impact": (
            "Inaccurate timekeeping between network devices can lead to many undesirable results "
            "including: inability to correlate events between devices making debugging difficult, "
            "failure of network authentication services that require time synchronization, and "
            "time-based network attacks that exploit time setting differences in network devices."
        ),
        "recommendation": (
            "It is recommended to configure NTP time services with at least two servers.\n\n"
            "switch(config)# clock datetime 2025-06-30 14:22:00\n"
            "switch(config)# clock timezone americas/los_angeles\n"
            "switch(config)# ntp server 10.10.10.10 iburst minpoll 10 maxpoll 14 prefer\n"
            "switch(config)# ntp server 20.20.20.20\n"
            "switch(config)# ntp vrf mgmt\n"
            "switch(config)# ntp enable"
        ),
        "expected_output": (
            "switch(config)# ntp server 10.10.10.10 iburst prefer\n"
            "switch(config)# ntp server 20.20.20.20\n"
            "switch(config)# ntp vrf mgmt\n"
            "switch(config)# ntp enable"
        ),
        "check_fn": "check_ntp_configured",
    },

    # ── 1.4 Securing SNMP Access ──────────────

    {
        "id": 11,
        "cis_id": "1.4.1.1",
        "obs_title": "Non Default Community Names, Access Rights & ACL (Automated)",
        "observation": "It was observed that SNMP is using default community names or no ACL is applied.",
        "severity": "High",
        "description": (
            "The default SNMP community name is 'public', a common setting for SNMP-capable devices. "
            "SNMPv1 and v2c use community names for read and write access. These community names are "
            "sent across the wire as clear text. If a malicious user were to capture these community "
            "names, they could potentially issue SNMP set commands to make unauthorized and potentially "
            "harmful configuration changes to a network device."
        ),
        "impact": (
            "By using a non-default (and strong) SNMP community name, you make it significantly "
            "harder for attackers to gain access to the SNMP interface of your devices. Without "
            "renaming the default community and applying ACLs, any host with network reachability "
            "can query or potentially modify device configuration via SNMP."
        ),
        "recommendation": (
            "It is recommended to replace default SNMP community names and apply ACLs.\n\n"
            "switch(config)# snmp-server community <community-name>\n"
            "switch(config-community)# access-level <ro | rw>\n"
            "switch(config-community)# access-list <IPv4 and/or IPv6 ACL>"
        ),
        "expected_output": (
            "switch(config)# snmp-server community <community-name>\n"
            "switch(config-community)# access-level ro\n"
            "switch(config-community)# access-list <ACL-NAME>"
        ),
        "check_fn": "check_snmp_community",
    },

    {
        "id": 12,
        "cis_id": "1.4.2.1",
        "obs_title": "SNMP V3 (Automated)",
        "observation": "It was observed that SNMPv3 is not configured or SNMPv3-only mode is not enforced.",
        "severity": "High",
        "description": (
            "SNMPv3 offers support for different users, authentication, and strong encryption. "
            "AOS-CX supports stronger authentication protocols (SHA224, SHA256, SHA384, and SHA512) "
            "and privacy protocols (AES192 and AES256). Enabling SNMPv3 enhances the security and "
            "reliability of network management by introducing authentication, encryption and message "
            "integrity features, addressing vulnerabilities present in earlier versions like "
            "SNMPv1 and SNMPv2c."
        ),
        "impact": (
            "Implementing SNMPv3 helps organizations meet security requirements and stay compliant "
            "with industry standards. SNMPv1 and v2c transmit community strings in clear text, "
            "exposing credentials to eavesdropping. SNMPv3 eliminates this risk through "
            "cryptographic authentication and privacy."
        ),
        "recommendation": (
            "It is recommended to configure SNMPv3 and disable support for SNMPv1/v2c.\n\n"
            "switch(config)# snmpv3 user <myUser> auth sha auth-pass plaintext <myAuthPswrd> "
            "priv des priv-pass plaintext <myPrivPswrd>\n"
            "switch(config)# snmp-server snmpv3-only\n"
            "switch(config)# snmp-server vrf <vrf-name>"
        ),
        "expected_output": (
            "switch(config)# snmpv3 user <myUser> auth sha auth-pass plaintext <myAuthPswrd> "
            "priv des priv-pass plaintext <myPrivPswrd>\n"
            "switch(config)# snmp-server snmpv3-only"
        ),
        "check_fn": "check_snmpv3",
    },

    {
        "id": 13,
        "cis_id": "1.4.3",
        "obs_title": "SNMP Traps (Automated)",
        "observation": "It was observed that SNMP traps are not fully configured on the device.",
        "severity": "Low",
        "description": (
            "Configuring SNMP traps in AOS-CX switches enables proactive monitoring and management "
            "capabilities within the network infrastructure. SNMP traps act as automated notifications "
            "sent from the switches to a designated management system whenever specific events or "
            "thresholds occur, such as high CPU utilization, port status changes, or AAA server "
            "unreachability. AOS-CX switches support multiple SNMP traps, which are either enabled "
            "by default or through configuration."
        ),
        "impact": (
            "Configuring SNMP traps in AOS-CX switches enables proactive issue detection, "
            "reducing downtime and improving response times. Without SNMP traps, administrators "
            "must rely on manual polling to detect issues, delaying response and potentially missing "
            "critical security events such as configuration changes or AAA server unreachability."
        ),
        "recommendation": (
            "It is recommended to configure SNMP traps for critical security events.\n\n"
            "switch(config)# snmp-server trap configuration-changes\n"
            "switch(config)# snmp-server trap aaa-server-reachability-status\n"
            "switch(config)# snmp-server host <MGMT-SERVER> trap version v3 user <WORD>"
        ),
        "expected_output": (
            "switch(config)# snmp-server trap configuration-changes\n"
            "switch(config)# snmp-server trap aaa-server-reachability-status\n"
            "switch(config)# snmp-server host <MGMT-SERVER> trap version v3 user <WORD>"
        ),
        "check_fn": "check_snmp_traps",
    },

    # ── 1.5 AAA for Management Interfaces ─────

    {
        "id": 14,
        "cis_id": "1.5.1.1",
        "obs_title": "Radius Server Configuration (Automated)",
        "observation": "It was observed that a RADIUS server is not configured for remote AAA.",
        "severity": "Medium",
        "description": (
            "Remote AAA with RADIUS provides authentication using remote RADIUS AAA servers and "
            "transmission of locally collected accounting information to remote RADIUS servers. "
            "Using a RADIUS server for AAA centralizes user management, ensuring uniform access "
            "policies across the network. It provides enhanced scalability, making it suitable "
            "for large networks with numerous users, compared to local AAA which is limited to "
            "individual switches."
        ),
        "impact": (
            "Implementing RADIUS server for AAA enhances security by minimizing the attack surface "
            "and ensuring compliance with industry standards for authentication. Without centralized "
            "AAA, each device must manage its own credentials, increasing the risk of inconsistent "
            "password policies and making credential rotation difficult."
        ),
        "recommendation": (
            "It is recommended to configure a RADIUS server for centralized AAA.\n\n"
            "switch(config)# radius-server host {<FQDN> | <IPV4> | <IPV6>} [key plaintext "
            "<PASSKEY>] [vrf <VRF-NAME>]\n"
            "switch(config)# aaa group server radius <group-name>\n"
            "switch(config-sg)# server <FQDN | IPv4 | IPv6> vrf <vrf>\n"
            "switch(config-sg)# exit"
        ),
        "expected_output": (
            "switch(config)# radius-server host <IP> key plaintext <KEY> vrf mgmt\n"
            "switch(config)# aaa group server radius <group-name>\n"
            "switch(config-sg)# server <IP> vrf mgmt"
        ),
        "check_fn": "check_radius_server",
    },

    {
        "id": 15,
        "cis_id": "1.5.1.2",
        "obs_title": "TACACS Server Configuration (Automated)",
        "observation": "It was observed that a TACACS+ server is not configured for remote AAA.",
        "severity": "Medium",
        "description": (
            "TACACS+ is a centralized authentication, authorization, and accounting protocol that "
            "provides granular control over user access and command-level authorization for AOS-CX "
            "switches. It separates authentication, authorization, and accounting processes for "
            "enhanced flexibility. TACACS+ supports per-command authorization, making it ideal for "
            "switches where granular control is needed. Unlike RADIUS, it encrypts entire packets, "
            "ensuring better security."
        ),
        "impact": (
            "Implementing TACACS+ ensures improved security and centralized control, reducing the "
            "risk of unauthorized access or misconfiguration. It simplifies user management and "
            "audit trails, enhancing operational efficiency and compliance. Local authentication "
            "lacks scalability and centralized management, making TACACS+ preferable for larger "
            "networks."
        ),
        "recommendation": (
            "It is recommended to configure a TACACS+ server for centralized AAA.\n\n"
            "switch(config)# tacacs-server host {<FQDN> | <IPV4> | <IPV6>} [key plaintext "
            "<PASSKEY>] [vrf <VRF-NAME>]\n"
            "switch(config)# aaa group server tacacs <group-name>\n"
            "switch(config-sg)# server <FQDN | IPv4 | IPv6> vrf <vrf>\n"
            "switch(config-sg)# exit"
        ),
        "expected_output": (
            "switch(config)# tacacs-server host <IP> key plaintext <KEY> vrf mgmt\n"
            "switch(config)# aaa group server tacacs <group-name>\n"
            "switch(config-sg)# server <IP> vrf mgmt"
        ),
        "check_fn": "check_tacacs_server",
    },

    {
        "id": 16,
        "cis_id": "1.5.2.2",
        "obs_title": "Limit Login Attempts (Automated)",
        "observation": "It was observed that failed login attempt limits are not configured for authentication profiles.",
        "severity": "Medium",
        "description": (
            "The 'Limiting Login Attempts' feature on Aruba AOS-CX switches is designed to "
            "enhance the security of the system by restricting the number of consecutive failed "
            "login attempts allowed for user accounts. Once the maximum number of failed attempts "
            "is reached, the account is blocked for a predetermined period - lockout-time. This "
            "local login attempt limiting feature is only available when not using remote "
            "authentication through AAA servers on any interface."
        ),
        "impact": (
            "Limiting login attempts on AOS-CX switches helps prevent brute-force attacks and "
            "unauthorized access, protecting critical network resources. This enhances security by "
            "safeguarding administrative accounts, reduces operational risks, and ensures compliance "
            "with security best practices."
        ),
        "recommendation": (
            "It is recommended to enable login attempt failure limiting with a 60 second lockout.\n\n"
            "For the console interface only:\n"
            "switch(config)# aaa authentication console-login-attempts 3 console-lockout-time 60\n\n"
            "For SSH/Telnet/https-server:\n"
            "switch(config)# aaa authentication limit-login-attempts 3 lockout-time 60"
        ),
        "expected_output": (
            "switch(config)# aaa authentication console-login-attempts 3 console-lockout-time 60\n"
            "switch(config)# aaa authentication limit-login-attempts 3 lockout-time 60"
        ),
        "check_fn": "check_failed_attempts",
    },

    {
        "id": 17,
        "cis_id": "1.5.2.3",
        "obs_title": "Remote Authentication - RADIUS/RadSec/TACACS+ (Automated)",
        "observation": "It was observed that remote authentication is not configured for management interfaces.",
        "severity": "Medium",
        "description": (
            "Remote Authentication involves the use of remote RADIUS, RadSec, and TACACS+ servers "
            "for authenticating management users. Remote AAA servers are used as a single point of "
            "management to configure and store user accounts. They are often coupled with directories "
            "and management repositories, simplifying the setup and maintenance of the end-user "
            "accounts. Remote authentication eliminates the need to manage local user accounts on "
            "each AOS-CX switch."
        ),
        "impact": (
            "Without remote authentication, each switch must independently manage user credentials, "
            "leading to inconsistent access controls across the network. Centralized remote "
            "authentication allows for uniform enforcement of security policies and ensures that "
            "all switches adhere to the same authentication and access standards."
        ),
        "recommendation": (
            "It is recommended to enable remote authentication for management interfaces.\n\n"
            "switch(config)# aaa authentication login <console/default/ssh/telnet/https-server> "
            "<RADIUS/TACACS+ group-list>"
        ),
        "expected_output": (
            "switch(config)# aaa authentication login default <group-list> local"
        ),
        "check_fn": "check_remote_auth",
    },

    {
        "id": 18,
        "cis_id": "1.5.3.1",
        "obs_title": "Local Authorization (Automated)",
        "observation": "It was observed that local authorization is not configured for management interfaces.",
        "severity": "Low",
        "description": (
            "Authorization controls authenticated users' command execution and switch interaction "
            "privileges. Local authorization uses role-based access control (RBAC) to provide "
            "role-based privilege levels plus optional user-defined local user groups with command "
            "execution rules. Authorization occurs only after successful authentication. "
            "Administrators have full command execution privilege, Operators are limited to "
            "non-sensitive show commands, and Auditors are limited to auditing-related commands."
        ),
        "impact": (
            "Using local authorization on CX switches ensures faster authorization processes and "
            "eliminates dependency on external servers, enhancing reliability in scenarios where "
            "network connectivity to remote servers may be unavailable. Without defined authorization, "
            "all authenticated users may gain full administrative access."
        ),
        "recommendation": (
            "It is recommended to enable local as the primary or fallback authorization method.\n\n"
            "switch(config)# aaa authorization commands {default | console | ssh | telnet} local"
        ),
        "expected_output": (
            "switch(config)# aaa authorization commands default local"
        ),
        "check_fn": "check_local_authorization",
    },

    {
        "id": 19,
        "cis_id": "1.5.3.2",
        "obs_title": "Remote Authorization (Automated)",
        "observation": "It was observed that remote authorization via TACACS+ is not configured.",
        "severity": "Medium",
        "description": (
            "Remote authorization on CX switches leverages centralized servers like TACACS+ to "
            "manage and authorize user access, ensuring consistent enforcement of policies across "
            "the network. Centralized remote authorization simplifies access management, enhances "
            "scalability, and provides a unified approach to monitoring and auditing user activities, "
            "reducing administrative overhead."
        ),
        "impact": (
            "This method strengthens security by centralizing control, ensures compliance with "
            "organizational policies, and allows for dynamic updates to access permissions, "
            "improving overall network governance and flexibility."
        ),
        "recommendation": (
            "It is recommended to configure remote authorization via TACACS+ with local fallback.\n\n"
            "switch(config)# aaa authorization commands {default | console | ssh | telnet} "
            "group <tacacs-server-group-list> local"
        ),
        "expected_output": (
            "switch(config)# aaa authorization commands default group <tacacs-group> local"
        ),
        "check_fn": "check_remote_authorization",
    },

    {
        "id": 20,
        "cis_id": "1.5.4.1",
        "obs_title": "Local Accounting (Automated)",
        "observation": "It was observed that local accounting is not configured on the device.",
        "severity": "Low",
        "description": (
            "Local Accounting records all CLI and REST access activities by users from all channels. "
            "It logs and helps to track all configuration changes and show command executions at the "
            "switch for auditing or accounting purposes. This includes: Exec Accounting (user "
            "login/logout events), Command accounting (commands executed by users), System accounting "
            "(remote accounting On/Off events), and interactions on non-CLI interfaces such as REST "
            "and WebUI."
        ),
        "impact": (
            "Local accounting logs provide immediate access to user activity and configuration changes "
            "directly on the switch, ensuring visibility even when external logging systems are "
            "unavailable. This enhances operational resilience and supports faster troubleshooting "
            "during network outages or remote accounting server failures."
        ),
        "recommendation": (
            "It is recommended to enable local accounting for all management actions.\n\n"
            "switch(config)# aaa accounting commands {exec|command|system} local"
        ),
        "expected_output": (
            "switch(config)# aaa accounting commands exec local\n"
            "switch(config)# aaa accounting commands command local"
        ),
        "check_fn": "check_local_accounting",
    },

    {
        "id": 21,
        "cis_id": "1.5.4.2",
        "obs_title": "Remote Accounting (Automated)",
        "observation": "It was observed that remote accounting is not configured for management interfaces.",
        "severity": "Medium",
        "description": (
            "For remote accounting, the information is sent to the first reachable remote server "
            "configured with this command. If no remote server is reachable, local accounting "
            "remains available by default. Implementing remote accounting enhances operational "
            "efficiency by providing real-time insights into network performance and security. "
            "It reduces the need for on-site management and ensures compliance with regulatory "
            "requirements through detailed reporting."
        ),
        "impact": (
            "Remote accounting improves network scalability and security, minimizes downtime "
            "through proactive monitoring, and reduces operational costs by enabling remote "
            "management of CX switches. Without remote accounting, security events may only be "
            "retained locally and can be lost if the device is rebooted or storage fills up."
        ),
        "recommendation": (
            "It is recommended to configure remote accounting to a centralized server.\n\n"
            "switch(config)# aaa accounting all-mgmt <console|default|https-server|ssh|telnet> "
            "start-stop <group>"
        ),
        "expected_output": (
            "switch(config)# aaa accounting all-mgmt default start-stop <tacacs-group>"
        ),
        "check_fn": "check_remote_accounting",
    },

    {
        "id": 22,
        "cis_id": "1.5.6",
        "obs_title": "Login Privilege Elevation for Administrators (Automated)",
        "observation": "It was observed that login privilege elevation is not enabled for administrators.",
        "severity": "Medium",
        "description": (
            "By default, the AOS-CX switches only allow the administrator user to perform the "
            "'enable' command and when executed, the user is elevated in privilege without "
            "prompting for a password. To increase security, the login privilege elevation feature "
            "can be enabled. Upon successful login, the administrator user will first be provided "
            "with lower privilege access (operator level '>' prompt). Upon executing the 'enable' "
            "command, the user will be prompted to enter a password for re-authentication before "
            "gaining administrator access ('#' prompt)."
        ),
        "impact": (
            "The impact of this feature enhances security and operational efficiency by ensuring "
            "that only trusted users gain administrative access after executing the 'enable' CLI. "
            "Every time a user runs the 'enable' CLI, highly secure environments advise demanding "
            "a password to prevent unauthorized privilege escalation."
        ),
        "recommendation": (
            "It is recommended to enable login privilege elevation for administrator group.\n\n"
            "switch(config)# aaa authentication login privilege-elevation group administrators"
        ),
        "expected_output": (
            "switch(config)# aaa authentication login privilege-elevation group administrators"
        ),
        "check_fn": "check_privilege_elevation",
    },

    # ── 1.6 PKI ───────────────────────────────

    {
        "id": 23,
        "cis_id": "1.6.2",
        "obs_title": "TLS Minimum Version (Automated)",
        "observation": "It was observed that TLS minimum version is not enforced on the device.",
        "severity": "Medium",
        "description": (
            "The switch supports TLS version 1.2 and 1.3 and secure cipher suites for each TLS "
            "version supported. Enforcing a minimum TLS version ensures that legacy, insecure "
            "TLS versions (1.0, 1.1) are not accepted. TLS 1.2 is the minimum recommended version "
            "as older versions contain known vulnerabilities that can be exploited."
        ),
        "impact": (
            "Allowing older TLS versions exposes management traffic to known vulnerabilities such "
            "as POODLE, BEAST, and CRIME attacks. Enforcing TLS 1.2 as minimum ensures all "
            "management plane communications use strong cryptographic standards, protecting data "
            "confidentiality and integrity in transit."
        ),
        "recommendation": (
            "It is recommended to enforce TLS 1.2 as the minimum version.\n\n"
            "switch(config)# tls minimum-version tls12"
        ),
        "expected_output": (
            "switch# show tls global\n"
            "Minimum TLS version : 1.2"
        ),
        "check_fn": "check_tls_version",
    },

    # ── 1.9 Securing https-server ─────────────

    {
        "id": 24,
        "cis_id": "1.9.1",
        "obs_title": "https-server default enablement (Automated)",
        "observation": "It was observed that the https-server is enabled on VRFs where it may not be required.",
        "severity": "Low",
        "description": (
            "On AOS-CX, the https-server is enabled by default allowing customers to access the "
            "device via its REST API or Web interface (WebUI). On campus products, the https-server "
            "is enabled on the default VRF. Products with an OOBM interface also have the https-server "
            "enabled on the mgmt VRF. Customers should be aware of this default enablement so that "
            "they can account for enabled management interfaces."
        ),
        "impact": (
            "Lack of knowledge of enabled management interfaces can result in unexpected access "
            "of a network device. If the https-server is not required for management purposes, "
            "disabling it reduces the attack surface and eliminates unnecessary access vectors "
            "to the device's management plane."
        ),
        "recommendation": (
            "It is recommended to disable the https-server on VRFs where it is not required.\n\n"
            "switch(config)# no https-server vrf <VRF-NAME>"
        ),
        "expected_output": (
            "switch# show https-server\n"
            "HTTPS Server Configuration\n"
            " VRF                 : mgmt"
        ),
        "check_fn": "check_https_server",
    },

    {
        "id": 25,
        "cis_id": "1.9.2",
        "obs_title": "https-server idle session management (Automated)",
        "observation": "It was observed that https-server idle session timeout is not configured to a secure value.",
        "severity": "Low",
        "description": (
            "https-server sessions are used by both the REST API and Web management interface "
            "(WebUI). By default, an https-server session will remain open for 20 minutes while "
            "idle. Customers should be setting the idle session timeout to the lowest value that "
            "can be tolerated. If only the WebUI is being used, an idle timeout of 5 minutes is "
            "recommended."
        ),
        "impact": (
            "Sessions that are kept open too long may be taken over if the administrator walks "
            "away from their computer and fails to lock it. This can lead to someone else gaining "
            "access to the device through the open management session. Reducing the session timeout "
            "minimizes the window of opportunity for session hijacking."
        ),
        "recommendation": (
            "It is recommended to configure the https-server idle session timeout to 5 minutes or less.\n\n"
            "switch(config)# https-server session-timeout 5"
        ),
        "expected_output": (
            "switch# show https-server\n"
            " Session timeout           : 5"
        ),
        "check_fn": "check_https_timeout",
    },

    # ── 1.10 ServiceOS Password ────────────────

    {
        "id": 26,
        "cis_id": "1.10.1",
        "obs_title": "ServiceOS Password (Automated)",
        "observation": "It was observed that ServiceOS password authentication is not enabled.",
        "severity": "Medium",
        "description": (
            "By default, the ServiceOS shell (accessible only from the local switch console port) "
            "requires no password to login as admin. This will allow un-privileged users with "
            "console access to log in to the ServiceOS shell. Enabling ServiceOS password "
            "authentication ensures only privileged users can access the ServiceOS shell, "
            "which provides low-level access to switch hardware and firmware."
        ),
        "impact": (
            "Enabling ServiceOS password authentication prevents unintended users from logging "
            "into the ServiceOS shell, significantly reducing the risk of various security "
            "vulnerabilities and attacks. Without this protection, any user with physical console "
            "access can access the ServiceOS shell with full admin privileges."
        ),
        "recommendation": (
            "It is recommended to enable ServiceOS password authentication.\n\n"
            "switch(config)# system serviceos password-prompt"
        ),
        "expected_output": (
            "switch(config)# show system serviceos password-prompt\n"
            "ServiceOS Password Prompt: Enabled"
        ),
        "check_fn": "check_serviceos_password",
    },

    # ── 1.11 Securing Syslog ──────────────────

    {
        "id": 27,
        "cis_id": "1.11.2",
        "obs_title": "Configure syslog-client to log using TLS (Automated)",
        "observation": "It was observed that logging to a remote syslog server using TLS is not configured.",
        "severity": "Low",
        "description": (
            "Logging to a remote syslog server should be configured to use TLS and include "
            "auditable events. Logging to a remote syslog server offers several advantages, "
            "including centralized log management, improved security, long-term data retention, "
            "and enhanced troubleshooting capabilities. By consolidating logs from various devices "
            "into a single location, administrators can easily monitor, analyze, and respond to "
            "issues across the network."
        ),
        "impact": (
            "Without a centralized logging location, logs on the switch may roll over and be lost. "
            "Security-relevant events including authentication attempts, configuration changes, "
            "privilege escalations, and network anomalies are stored only locally and can be "
            "lost if the device is rebooted, cleared, or compromised."
        ),
        "recommendation": (
            "It is recommended to configure the switch to send logs to a syslog-server using TLS.\n\n"
            "switch(config)# logging <SYSLOG-SERVER> tls <PORT> auth-mode subject-name vrf mgmt "
            "include-auditable-events"
        ),
        "expected_output": (
            "switch(config)# logging <SYSLOG-SERVER> tls <PORT> auth-mode subject-name vrf mgmt "
            "include-auditable-events"
        ),
        "check_fn": "check_syslog",
    },

    # ── 1.12 Login Banner ─────────────────────

    {
        "id": 28,
        "cis_id": "1.12",
        "obs_title": "Login Banner (Automated)",
        "observation": "It was observed that a login banner (MOTD/exec) is not configured on the device.",
        "severity": "Low",
        "description": (
            "Setting a banner to be displayed during the login process notifies users that "
            "unauthorized use is prohibited, and that access to and use of the system may be "
            "monitored and logged. Login Banner serves as a compliance measure with legal "
            "requirements, alerting users about usage terms or unauthorized access penalties. "
            "It enhances security awareness, reinforces accountability, and ensures clarity in "
            "user interactions with the network device."
        ),
        "impact": (
            "Without a login banner, there is no legal warning presented to users before accessing "
            "the device. This may affect the organization's ability to prosecute unauthorized access "
            "and may fail compliance requirements that mandate warning banners on all network devices."
        ),
        "recommendation": (
            "It is recommended to configure both pre-login and post-login banners.\n\n"
            "To configure Message of the Day (Pre-Login) Banner:\n"
            "switch(config)# banner motd ^\n"
            "switch(config-banner-motd)# This system is for authorized use only. ^\n\n"
            "To configure Post-Login Banner:\n"
            "switch(config)# banner exec ^\n"
            "switch(config-banner-exec)# This banner is displayed after login ^\n"
        ),
        "expected_output": (
            "switch(config)# banner motd ^\n"
            "switch(config-banner-motd)# <Banner Text> ^\n"
            "switch(config)# banner exec ^\n"
            "switch(config-banner-exec)# <Banner Text> ^"
        ),
        "check_fn": "check_login_banner",
    },

    # ── 1.13 Configuration Backup ─────────────

    {
        "id": 29,
        "cis_id": "1.13",
        "obs_title": "Schedule Configuration Backup Job (Automated)",
        "observation": "It was observed that no scheduled configuration backup job is configured.",
        "severity": "Low",
        "description": (
            "Creating a job to periodically backup the configuration to a remote server is crucial "
            "for disaster recovery and business continuity. Remote backups protect against local "
            "disasters, cyberattacks including ransomware, hardware failures, and human errors. "
            "Automated backups reduce manual effort and minimize the risk of human error. Remote "
            "backup solutions also support scalability and compliance requirements for data retention."
        ),
        "impact": (
            "If no remote configuration backup strategy is implemented, the organization is exposed "
            "to issues with business continuity, disaster recovery, and compliance. Loss of the "
            "device configuration could result in significant downtime and recovery effort following "
            "any failure or security incident."
        ),
        "recommendation": (
            "It is recommended to create and schedule a periodic configuration backup job.\n\n"
            "switch(config)# job backup_config\n"
            "switch(config-job-config_backup)# 1 cli copy running-config "
            "sftp://root@<SERVER>/switch.cfg cli vrf mgmt\n"
            "switch(config)# schedule backup_config\n"
            "switch(config-scheduler-backup_config)# 1 job backup_config\n"
            "switch(config-scheduler-backup_config)# trigger every days 1 start 12:00 2025-01-01\n"
            "switch(config-scheduler-backup_config)# enable"
        ),
        "expected_output": (
            "switch(config)# job backup_config\n"
            "switch(config)# schedule backup_config\n"
            "switch(config-scheduler-backup_config)# enable"
        ),
        "check_fn": "check_config_backup",
    },

    # ── 1.14 Hostname ─────────────────────────

    {
        "id": 30,
        "cis_id": "1.14",
        "obs_title": "Create Hostname (Automated)",
        "observation": "It was observed that the device does not have a unique hostname configured.",
        "severity": "Low",
        "description": (
            "The hostname property should be set to a unique name on the management network. "
            "This enables the switch to be resolved with a unique FQDN. In networks with multiple "
            "switches, a unique hostname makes it easy to identify each device quickly and avoids "
            "confusion when managing logs, alerts, and configurations. Hostnames appear in CLI "
            "prompts, SNMP, syslog messages, and monitoring tools. Many organizations require "
            "meaningful hostnames for inventory and audit purposes."
        ),
        "impact": (
            "Without a unique and meaningful hostname, identifying the device in logs, SNMP traps, "
            "and syslog messages becomes difficult. This slows down incident response and "
            "troubleshooting, and may cause confusion during configuration management activities."
        ),
        "recommendation": (
            "It is recommended to configure a unique, descriptive hostname.\n\n"
            "switch(config)# hostname <unique-device-name>"
        ),
        "expected_output": (
            "switch(config)# hostname <unique-device-name>"
        ),
        "check_fn": "check_hostname",
    },

    # ══════════════════════════════════════════
    # SECTION 2 – INTERFACE
    # ══════════════════════════════════════════

    # ── 2.1 Physical Security ─────────────────

    {
        "id": 31,
        "cis_id": "2.1.1",
        "obs_title": "Disable USB and Bluetooth on Device (Automated)",
        "observation": "It was observed that USB or Bluetooth is enabled on the device.",
        "severity": "Medium",
        "description": (
            "The AOS-CX switch front-panel includes a USB Auxiliary port for USB Mass storage "
            "(flash drive for deploying, troubleshooting, backing up configurations, or upgrading "
            "switches) and Bluetooth Adapter (allows Bluetooth-enabled devices to connect to and "
            "manage the switch wirelessly). The Bluetooth feature has been enabled by default in "
            "AOS-CX switches. Paired devices can manage the switch through SSH, Web UI, REST API, "
            "or the Aruba CX Mobile App."
        ),
        "impact": (
            "Disabling USB will prevent both USB devices from being mounted and Bluetooth adapters "
            "from being enabled. Disabling Bluetooth will prevent Bluetooth adapters from being "
            "enabled, reducing the wireless attack surface. Without disabling these interfaces, "
            "an attacker with physical proximity could attempt to connect via Bluetooth or insert "
            "a USB device to access or modify the switch."
        ),
        "recommendation": (
            "It is recommended to disable USB and Bluetooth on the device.\n\n"
            "switch(config)# no usb\n"
            "switch(config)# bluetooth disable"
        ),
        "expected_output": (
            "switch(config)# no usb\n"
            "switch(config)# bluetooth disable"
        ),
        "check_fn": "check_usb_bluetooth",
    },

    {
        "id": 32,
        "cis_id": "2.1.3",
        "obs_title": "Disable Unused Physical Interfaces (Manual)",
        "observation": "It was observed that unused physical interfaces are not disabled on the device.",
        "severity": "Medium",
        "description": (
            "Unused physical interfaces could allow an attacker to plug in a network cable and "
            "access network resources depending on the configuration of that port. Disabling unused "
            "physical interfaces helps to reduce the attack surface by minimizing open entry points, "
            "prevent unauthorized network access, ensure compliance with regulatory standards that "
            "require only necessary network ports to be active, and improve network hygiene by "
            "regularly managing and configuring port settings."
        ),
        "impact": (
            "Active but unused interfaces provide open network access points that could be "
            "exploited by an attacker who connects a device to an unmonitored port. Disabling "
            "unused interfaces enhances incident response capability by making it easier to isolate "
            "and respond effectively to security breaches with fewer active ports."
        ),
        "recommendation": (
            "It is recommended to disable all unused physical interfaces.\n\n"
            "switch(config)# interface <interface-id>\n"
            "switch(config-if)# shutdown\n"
            "switch(config-if)# exit"
        ),
        "expected_output": (
            "switch(config)# interface <unused-interface-id>\n"
            "switch(config-if)# shutdown"
        ),
        "check_fn": "check_unused_interfaces",
    },

    {
        "id": 33,
        "cis_id": "2.2",
        "obs_title": "Traffic Control - Rate limiting (Manual)",
        "observation": "It was observed that rate limiting is not configured on interfaces.",
        "severity": "Low",
        "description": (
            "Configuring rate-limiting on Aruba CX switches allows administrators to control the "
            "maximum amount of traffic (bandwidth) that can pass through specific interfaces or for "
            "specific traffic types. This ensures that no single device or application can consume "
            "excessive network resources. Rate-limiting types include: broadcast, multicast, "
            "unknown-unicast, or ICMP, applied in kilobits per second, packets per second, or as "
            "a percentage of link bandwidth."
        ),
        "impact": (
            "Rate-limiting is implemented to prevent network congestion, mitigate the impact of "
            "potential denial-of-service (DoS) attacks, and ensure fair bandwidth allocation among "
            "users and applications. By configuring rate-limiting, organizations can protect critical "
            "services from being overwhelmed by excessive traffic and ensure a consistent user "
            "experience."
        ),
        "recommendation": (
            "It is recommended to configure rate limiting on interfaces for broadcast and ICMP traffic.\n\n"
            "switch(config)# interface <interface-ID>\n"
            "switch(config-if)# rate-limit broadcast 500 kbps\n"
            "switch(config-if)# rate-limit icmp ip 10000 kbps"
        ),
        "expected_output": (
            "switch(config)# interface <interface-ID>\n"
            "switch(config-if)# rate-limit broadcast <RATE> kbps\n"
            "switch(config-if)# rate-limit icmp ip <RATE> kbps"
        ),
        "check_fn": "check_rate_limiting",
    },

    {
        "id": 34,
        "cis_id": "2.3",
        "obs_title": "Proxy ARP (Manual)",
        "observation": "It was observed that Proxy ARP is enabled on one or more interfaces.",
        "severity": "Low",
        "description": (
            "Proxy ARP is a technique by which a device on a given network answers ARP queries for "
            "a network address that is not on that network. The ARP proxy is aware of the location "
            "of the traffic's destination, and offers its own MAC address as the final destination. "
            "Proxy ARP is supported on L3 physical and VLAN interfaces. It is disabled by default. "
            "Disabling Proxy ARP on unused interfaces prevents the switch from responding to ARP "
            "requests on behalf of other devices."
        ),
        "impact": (
            "Disabling Proxy ARP on unused interfaces minimizes attack surfaces, safeguards against "
            "ARP-based attacks, and contributes to an optimized and secure network environment by "
            "eliminating unnecessary overhead on inactive ports. Enabling Proxy ARP unnecessarily "
            "can lead to spoofing or unauthorized traffic redirection."
        ),
        "recommendation": (
            "It is recommended to disable Proxy ARP on all interfaces where it is not required.\n\n"
            "switch(config)# interface <ID>\n"
            "switch(config-if)# no ip proxy-arp\n"
            "switch(config-if)# no ip local-proxy-arp\n"
            "switch(config-if)# no ipv6 local-proxy-nd"
        ),
        "expected_output": (
            "switch(config)# interface <ID>\n"
            "switch(config-if)# no ip proxy-arp\n"
            "switch(config-if)# no ip local-proxy-arp"
        ),
        "check_fn": "check_proxy_arp",
    },

    {
        "id": 35,
        "cis_id": "2.4",
        "obs_title": "IP Directed Broadcast (Manual)",
        "observation": "It was observed that IP Directed Broadcast is enabled on one or more interfaces.",
        "severity": "Medium",
        "description": (
            "IP Directed Broadcast is a feature by which remote administration tasks such as backups "
            "and wake-on-LAN (WOL) can be achieved by sending directed broadcast packets for hosts "
            "and servers residing on different subnets. It is disabled by default. Disabling IP "
            "directed broadcast on interfaces prevents the switch from converting directed broadcast "
            "packets into Layer 2 broadcasts on a destination subnet, ensuring such traffic is "
            "dropped rather than propagated further into the network."
        ),
        "impact": (
            "IP directed broadcasts are often exploited in amplification attacks, such as Smurf "
            "attacks, which flood the network with broadcast traffic to disrupt operations. By "
            "disabling this feature on unused interfaces, network administrators can reduce the "
            "likelihood of these types of exploits. It also prevents unnecessary broadcast traffic "
            "from consuming resources on unused subnets."
        ),
        "recommendation": (
            "It is recommended to disable IP Directed Broadcast on all interfaces.\n\n"
            "switch(config)# interface <ID>\n"
            "switch(config-if)# no ip directed-broadcast"
        ),
        "expected_output": (
            "switch(config)# interface <ID>\n"
            "switch(config-if)# no ip directed-broadcast"
        ),
        "check_fn": "check_directed_broadcast",
    },

    # ══════════════════════════════════════════
    # SECTION 3 – PROTOCOL
    # ══════════════════════════════════════════

    # ── 3.1.1 OSPF ────────────────────────────

    {
        "id": 36,
        "cis_id": "3.1.1.1",
        "obs_title": "OSPF Passive Interfaces (Manual)",
        "observation": "It was observed that OSPF passive interfaces are not configured.",
        "severity": "Low",
        "description": (
            "The OSPF Passive Interface feature is used to prevent OSPF from forming neighbor "
            "relationships or exchanging Hello packets on specified interfaces, while still allowing "
            "the connected subnet to be advertised in the OSPF routing domain. This enhances network "
            "security by avoiding unauthorized OSPF adjacencies, reduces unnecessary control traffic, "
            "and optimizes resource utilization by limiting OSPF operations to only required "
            "interfaces. It is particularly useful for interfaces connected to end-user devices or "
            "untrusted networks."
        ),
        "impact": (
            "Without passive interfaces, OSPF will attempt to form neighbor relationships on all "
            "enabled interfaces, including those connected to end-user devices or untrusted networks. "
            "This increases the risk of unauthorized OSPF adjacencies, routing table manipulation, "
            "and unnecessary control traffic consumption on edge interfaces."
        ),
        "recommendation": (
            "It is recommended to configure all OSPF-enabled interfaces as passive by default, "
            "then selectively enable OSPF peering only on required interfaces.\n\n"
            "OSPF:\n"
            "switch(config)# router ospf <OSPF Process ID>\n"
            "switch(config-ospf-<Process ID>)# passive-interface default\n\n"
            "OSPFv3:\n"
            "switch(config)# router ospfv3 <OSPF Process ID>\n"
            "switch(config-ospfv3-<Process ID>)# passive-interface default"
        ),
        "expected_output": (
            "switch(config)# router ospf <Process ID>\n"
            "switch(config-ospf)# passive-interface default"
        ),
        "check_fn": "check_ospf_passive",
    },

    {
        "id": 37,
        "cis_id": "3.1.1.2",
        "obs_title": "OSPF Neighbor Authentication (Manual)",
        "observation": "It was observed that OSPF neighbor authentication is not configured.",
        "severity": "Medium",
        "description": (
            "The purpose of configuring OSPF neighbor authentication is to secure communication "
            "between OSPF routers and protect the integrity of routing information exchanged. By "
            "default, OSPF uses 'null' authentication, meaning no authentication is applied, leaving "
            "OSPF control packets vulnerable to unauthorized access, tampering, or injection of "
            "malicious routing information. Introducing cryptographic authentication methods such as "
            "HMAC-SHA-512 ensures that OSPF neighbors exchange routing information securely."
        ),
        "impact": (
            "Without OSPF neighbor authentication, unauthorized devices can potentially become OSPF "
            "neighbors and inject malicious routing information, leading to routing table poisoning, "
            "traffic redirection, or network disruptions. HMAC-SHA-512 provides strong protection "
            "against brute force and collision attacks."
        ),
        "recommendation": (
            "It is recommended to configure OSPF neighbor authentication using HMAC-SHA-512.\n\n"
            "switch(config)# keychain <keychain-name>\n"
            "switch(config-keychain)# key <ID>\n"
            "switch(config-keychain-key)# cryptographic-algorithm hmac-sha-512\n"
            "switch(config-keychain-key)# key-string plaintext <plaintext-key-string>\n"
            "switch(config)# interface <interface-ID>\n"
            "switch(config-if)# ip ospf authentication keychain\n"
            "switch(config-if)# ip ospf keychain <keychain-name>"
        ),
        "expected_output": (
            "switch(config)# keychain <keychain-name>\n"
            "switch(config-keychain)# key <ID>\n"
            "switch(config-keychain-key)# cryptographic-algorithm hmac-sha-512\n"
            "switch(config-if)# ip ospf authentication keychain"
        ),
        "check_fn": "check_ospf_auth",
    },

    {
        "id": 38,
        "cis_id": "3.1.1.3",
        "obs_title": "OSPFv3 Area Authentication and Encryption with IPsec (Manual)",
        "observation": "It was observed that OSPFv3 area authentication with IPsec is not configured.",
        "severity": "Medium",
        "description": (
            "OSPFv3 neighbors may use interface-level authentication. IPsec can be used to provide "
            "encryption, or authentication, or both for an entire OSPFv3 area, automatically applying "
            "the configured methods to all member interfaces. There are two IPsec encapsulation types: "
            "IPv6 authentication header (AH) which adds an IPv6 authentication header to OSPFv3 "
            "packets, and Encrypted Security Payload (ESP) which provides both authentication and "
            "encryption. AOS-CX recommends using AES over DES or 3DES as it is stronger."
        ),
        "impact": (
            "The use of IPsec for OSPFv3 authentication and encryption significantly strengthens the "
            "security posture of networks by safeguarding routing updates in transit. This ensures "
            "that only authorized routers can participate in OSPFv3 exchanges and that all routing "
            "information remains confidential and tamper-proof, reducing risks of routing table "
            "manipulation, traffic redirection, and denial-of-service attacks."
        ),
        "recommendation": (
            "It is recommended to configure IPsec authentication for OSPFv3 areas.\n\n"
            "switch(config)# router ospfv3 <process-ID>\n"
            "switch(config-ospfv3-<process>)# area 1 authentication ipsec spi 1024 sha1\n"
            "Enter the IPsec authentication key: *******"
        ),
        "expected_output": (
            "switch(config)# router ospfv3 <process-ID>\n"
            "switch(config-ospfv3)# area <area-id> authentication ipsec spi <SPI> sha1"
        ),
        "check_fn": "check_ospfv3_ipsec",
    },

    # ── 3.1.2 BGP ─────────────────────────────

    {
        "id": 39,
        "cis_id": "3.1.2.1",
        "obs_title": "Control Plane ACL for BGP Peering Sessions (Manual)",
        "observation": "It was observed that a control plane ACL for BGP peering sessions is not configured.",
        "severity": "Medium",
        "description": (
            "Utilizing the control-plane ACL functionality to limit BGP communication to configured "
            "BGP peers. Control Plane ACLs (CP-ACLs) are essential for securing BGP peering sessions "
            "by restricting unauthorized access to the control plane. They prevent threats like "
            "spoofed packets, unauthorized session attempts, and malicious traffic, ensuring only "
            "legitimate BGP traffic is processed."
        ),
        "impact": (
            "Implementing CP-ACLs enhances network security and stability by safeguarding the "
            "control plane from attacks such as DDoS or unauthorized BGP manipulation. Without "
            "CP-ACLs, any host with IP reachability to the device can attempt to establish BGP "
            "sessions or flood the control plane with TCP port 179 traffic."
        ),
        "recommendation": (
            "It is recommended to configure control plane ACLs to restrict BGP peering to authorized peers.\n\n"
            "switch(config)# access-list ip <CONTROLPLANE-ACL-NAME>\n"
            "switch(config-acl-ip)# 805 permit tcp <peer-ip> gt 1023 any eq 179\n"
            "switch(config-acl-ip)# 810 permit tcp <peer-ip> eq 179 any gt 1023\n"
            "switch(config-acl-ip)# 815 deny tcp any gt 1023 any eq 179\n"
            "switch(config-acl-ip)# 1000 permit any any any\n"
            "switch(config)# apply access-list ip <CONTROLPLANE-ACL-NAME> control-plane vrf <VRF-NAME>"
        ),
        "expected_output": (
            "switch(config)# access-list ip <ACL-NAME>\n"
            "switch(config)# apply access-list ip <ACL-NAME> control-plane vrf <VRF-NAME>"
        ),
        "check_fn": "check_bgp_cp_acl",
    },

    {
        "id": 40,
        "cis_id": "3.1.2.2",
        "obs_title": "Authenticate BGP Peers Using MD5 (Manual)",
        "observation": "It was observed that BGP peer MD5 authentication is not configured.",
        "severity": "Medium",
        "description": (
            "The TCP sessions between two BGP peers can be secured by adding MD5 protection to the "
            "TCP session header. The MD5 digest acts like a password between peers. Authenticating "
            "BGP peers using MD5 ensures that only trusted and authorized devices can establish BGP "
            "sessions, protecting the integrity of the routing infrastructure. By adding a "
            "cryptographic layer of authentication, MD5 mitigates risks such as session hijacking "
            "and spoofed connection attempts."
        ),
        "impact": (
            "Using MD5 for BGP peer authentication significantly improves network security by "
            "preventing unauthorized devices from establishing BGP sessions. This reduces the risk "
            "of malicious activity such as route injection or disruption of routing operations, "
            "ensuring reliable and secure communication between peers."
        ),
        "recommendation": (
            "It is recommended to configure MD5 authentication for all BGP peers.\n\n"
            "switch(config)# router bgp <ASN>\n"
            "switch(config-bgp)# neighbor {<IP-ADDR>|<PEER-GROUP-NAME>} password "
            "[{ciphertext | plaintext} <PASSWORD>]"
        ),
        "expected_output": (
            "switch(config)# router bgp <ASN>\n"
            "switch(config-bgp)# neighbor <peer-ip> password plaintext <PASSWORD>"
        ),
        "check_fn": "check_bgp_md5",
    },

    {
        "id": 41,
        "cis_id": "3.1.2.3",
        "obs_title": "BGP TTL Security (Manual)",
        "observation": "It was observed that BGP TTL security is not configured for BGP peers.",
        "severity": "Low",
        "description": (
            "BGP TTL Security is designed to prevent unauthorized devices or attackers from "
            "establishing BGP sessions by leveraging the Time-to-Live (TTL) field in IP headers. "
            "By requiring BGP packets to have a specific TTL value (typically set high, e.g., 255), "
            "it ensures that only directly connected peers can communicate, blocking spoofed or "
            "malicious BGP traffic from distant sources."
        ),
        "impact": (
            "Implementing BGP TTL Security enhances the protection of BGP sessions by preventing "
            "attacks such as spoofed route advertisements or unauthorized session establishment "
            "from non-directly connected sources. This strengthens the network's resilience against "
            "threats and reduces the risk of routing instability."
        ),
        "recommendation": (
            "It is recommended to configure BGP TTL security for external BGP peers.\n\n"
            "switch(config)# router bgp <ASN>\n"
            "switch(config-bgp)# neighbor {<IP-ADDR>|<PEER-GROUP-NAME>} ttl-security-hops <hop-count>"
        ),
        "expected_output": (
            "switch(config)# router bgp <ASN>\n"
            "switch(config-bgp)# neighbor <peer-ip> ttl-security-hops <hop-count>"
        ),
        "check_fn": "check_bgp_ttl",
    },

    # ── 3.2 DHCP Protocols ────────────────────

    {
        "id": 42,
        "cis_id": "3.2.1.1",
        "obs_title": "DHCPv4 & DHCPv6 Snooping Enablement (Manual)",
        "observation": "It was observed that DHCP snooping is not enabled on the device.",
        "severity": "Medium",
        "description": (
            "DHCP snooping protects the network from common DHCP attacks, including address spoofing "
            "resulting from a rogue DHCP server operating on the network or exhaustion of addresses "
            "on a DHCP server caused by mass address requests. DHCP snooping designates trusted DHCP "
            "servers and ports on which DHCP requests and responses are accepted. DHCP packets are "
            "forwarded between trusted ports without inspection; packets received from untrusted "
            "sources are dropped."
        ),
        "impact": (
            "Without DHCP snooping, rogue DHCP servers can assign incorrect or malicious IP "
            "configuration to clients, leading to DHCP-based man-in-the-middle attacks, service "
            "disruption, and unauthorized network access. DHCP snooping prevents rogue DHCP servers "
            "from distributing unauthorized IP addresses and configurations."
        ),
        "recommendation": (
            "It is recommended to enable DHCP snooping globally and on all relevant VLANs.\n\n"
            "switch(config)# dhcpv4-snooping\n"
            "switch(config)# vlan <vlan-id>\n"
            "switch(config-vlan-id)# dhcpv4-snooping\n"
            "switch(config)# interface <trusted interface id>\n"
            "switch(config-if)# dhcpv4-snooping trust\n\n"
            "switch(config)# dhcpv6-snooping\n"
            "switch(config)# vlan <vlan-id>\n"
            "switch(config-vlan-id)# dhcpv6-snooping\n"
            "switch(config)# interface <trusted interface id>\n"
            "switch(config-if)# dhcpv6-snooping trust"
        ),
        "expected_output": (
            "switch(config)# dhcpv4-snooping\n"
            "switch(config-vlan-id)# dhcpv4-snooping\n"
            "switch(config-if)# dhcpv4-snooping trust"
        ),
        "check_fn": "check_dhcp_snooping",
    },

    {
        "id": 43,
        "cis_id": "3.2.1.2",
        "obs_title": "DHCPv6 Guard (Manual)",
        "observation": "It was observed that DHCPv6 Guard is not configured on the device.",
        "severity": "Low",
        "description": (
            "DHCPv6 guard is an extension of DHCPv6 snooping. When DHCPv6 snooping is configured "
            "globally and on the VLAN, DHCPv6 guard enhances this by creating a policy applied on "
            "a port and on the VLAN. This policy contains multiple attributes which are compared "
            "against packets received on trusted ports. If the packet complies with the attributes "
            "of the policy, it is forwarded; otherwise the packet is dropped."
        ),
        "impact": (
            "By mitigating the risks of rogue DHCPv6 servers, DHCPv6 Guard enhances network "
            "security, reduces potential misconfigurations, and ensures reliable IPv6 address "
            "management, leading to a more stable and secure network infrastructure."
        ),
        "recommendation": (
            "It is recommended to configure DHCPv6 guard policies on all relevant VLANs.\n\n"
            "switch(config)# dhcpv6-snooping guard-policy <policy-name>\n"
            "switch(config-dhcpv6-guard-policy)# match server access-list <acl-name>\n"
            "switch(config-dhcpv6-guard-policy)# exit\n"
            "switch(config)# vlan <vlan-id>\n"
            "switch(config-vlan-id)# dhcpv6-snooping guard-policy <policy-name>"
        ),
        "expected_output": (
            "switch(config)# dhcpv6-snooping guard-policy <policy-name>\n"
            "switch(config-vlan-id)# dhcpv6-snooping guard-policy <policy-name>"
        ),
        "check_fn": "check_dhcpv6_guard",
    },

    # ── 3.3 Multicast ─────────────────────────

    {
        "id": 44,
        "cis_id": "3.3.1.1",
        "obs_title": "PIM Accept-Register (Manual)",
        "observation": "It was observed that PIM Accept-Register is not configured.",
        "severity": "Low",
        "description": (
            "PIM Accept-register is a security feature designed to control which sources and groups "
            "are allowed to register with the Rendezvous Point (RP). By associating an ACL with a "
            "PIM router, administrators can filter and control the registration of multicast sources "
            "and groups. Without such control, unauthorized or malicious sources could register with "
            "the RP, potentially leading to resource misuse or network disruption."
        ),
        "impact": (
            "Packets matching the ACL permit rules are processed normally, while denied or unmatched "
            "packets are immediately dropped with a register stop message sent. This prevents "
            "unauthorized or malicious traffic from reaching the RP, enhancing network security "
            "and efficiency."
        ),
        "recommendation": (
            "It is recommended to configure PIM Accept-Register with an ACL to filter unauthorized sources.\n\n"
            "switch(config)# access-list ip <pim_reg_acl>\n"
            "switch(config-acl-ip)# 10 permit any <source> <group>\n"
            "switch(config-acl-ip)# exit\n"
            "switch(config)# router pim\n"
            "switch(config-pim)# accept-register access-list <pim_reg_acl>"
        ),
        "expected_output": (
            "switch(config)# router pim\n"
            "switch(config-pim)# accept-register access-list <pim_reg_acl>"
        ),
        "check_fn": "check_pim_accept_register",
    },

    {
        "id": 45,
        "cis_id": "3.3.1.2",
        "obs_title": "PIM Accept-RP (Manual)",
        "observation": "It was observed that PIM Accept-RP is not configured.",
        "severity": "Low",
        "description": (
            "PIM Accept-RP allows administrators to restrict which multicast groups a rendezvous "
            "point (RP) serves in a PIM sparse mode domain. By default, an RP accepts all multicast "
            "groups in the 224.0.0.0/4 range. Using an access list (ACL), this feature limits "
            "join/prune message processing to specific multicast groups, enhancing control and "
            "reducing unnecessary traffic."
        ),
        "impact": (
            "Without restrictions, an RP processes join/prune messages for all multicast groups in "
            "the default Class D range, leading to potential resource waste and security risks. "
            "PIM Accept-RP configured with an ACL ensures the RP handles only authorized multicast "
            "groups, preventing unwanted traffic and optimizing network operations."
        ),
        "recommendation": (
            "It is recommended to configure PIM Accept-RP with an ACL to restrict authorized groups.\n\n"
            "switch(config)# access-list ip <pim_rp_grp_acl>\n"
            "switch(config-acl-ip)# 10 permit any any <group/mask>\n"
            "switch(config-acl-ip)# exit\n"
            "switch(config)# router pim\n"
            "switch(config-pim)# accept-rp <rp-address> access-list <pim_rp_grp_acl>"
        ),
        "expected_output": (
            "switch(config)# router pim\n"
            "switch(config-pim)# accept-rp <rp-address> access-list <pim_rp_grp_acl>"
        ),
        "check_fn": "check_pim_accept_rp",
    },

    {
        "id": 46,
        "cis_id": "3.3.1.3",
        "obs_title": "PIM SSM (Manual)",
        "observation": "It was observed that PIM SSM range ACL is not configured.",
        "severity": "Low",
        "description": (
            "Protocol Independent Multicast - Source-Specific Multicast (PIM-SSM) is a specialized "
            "subset of PIM sparse mode that enables efficient delivery of multicast traffic from "
            "specific sources to receivers. By default, PIM-SSM operates within the group range "
            "232.0.0.0/8 for IPv4 and FF3x::/32 for IPv6. The feature also allows administrators "
            "to customize this range using an ACL to define valid multicast group addresses."
        ),
        "impact": (
            "PIM-SSM significantly improves multicast traffic efficiency by eliminating unnecessary "
            "traffic from unwanted sources, reducing network congestion, and ensuring only relevant "
            "multicast streams are delivered to receivers. Ensuring uniformity in the SSM range "
            "across a network is critical to maintaining seamless multicast traffic delivery."
        ),
        "recommendation": (
            "It is recommended to configure PIM-SSM with an ACL to define the authorized SSM range.\n\n"
            "switch(config)# access-list ip <pim_ssm_grp_range_acl>\n"
            "switch(config-acl-ip)# 10 permit any any <group-range>\n"
            "switch(config-acl-ip)# exit\n"
            "switch(config)# router pim\n"
            "switch(config-pim)# pim-ssm range-access-list <pim_ssm_grp_range_acl>"
        ),
        "expected_output": (
            "switch(config)# router pim\n"
            "switch(config-pim)# pim-ssm range-access-list <pim_ssm_grp_range_acl>"
        ),
        "check_fn": "check_pim_ssm",
    },

    {
        "id": 47,
        "cis_id": "3.3.3",
        "obs_title": "IGMP Snooping ACL (Manual)",
        "observation": "It was observed that an ACL is not applied to IGMP snooping.",
        "severity": "Low",
        "description": (
            "IGMP snooping operates on a Layer 2 device as a multicast constraining mechanism. "
            "The device may encounter invalid multicast entries if malicious users send IGMP reports, "
            "which can disrupt multicast services for legitimate users. Configuring a multicast group "
            "policy using an ACL allows the Layer 2 device to control which multicast groups hosts "
            "can join. The device filters IGMP reports according to the policy and adds the host's "
            "port to the outgoing port list only if the report is permitted by the policy."
        ),
        "impact": (
            "Without an IGMP snooping ACL, malicious users can send unauthorized IGMP reports to "
            "consume multicast hardware and software resources, disrupting legitimate multicast "
            "services. The ACL ensures that multicast traffic flows only to authorized groups, "
            "providing better security and optimized resource utilization."
        ),
        "recommendation": (
            "It is recommended to apply an ACL to IGMP snooping to filter unauthorized multicast groups.\n\n"
            "switch(config)# ip igmp snooping apply access list <ACL-NAME>"
        ),
        "expected_output": (
            "switch(config)# ip igmp snooping apply access list <ACL-NAME>"
        ),
        "check_fn": "check_igmp_acl",
    },

    {
        "id": 48,
        "cis_id": "3.3.4",
        "obs_title": "MLD Snooping ACL (Manual)",
        "observation": "It was observed that an ACL is not applied to MLD snooping.",
        "severity": "Low",
        "description": (
            "Multicast Listener Discovery (MLD) snooping is a feature designed to optimize multicast "
            "traffic across the network by preventing multicast flooding within a VLAN. Similar to "
            "IGMP, malicious users can generate invalid multicast entries using MLD reports, which "
            "can overwhelm the device and disrupt legitimate multicast services. A multicast group "
            "policy configured with an ACL is used to filter MLD reports."
        ),
        "impact": (
            "When an ACL is applied to MLD snooping, only permitted group addresses in MLD packets "
            "are processed. Ports that do not match the MLD group policy are blocked from receiving "
            "multicast traffic, ensuring that unauthorized or excessive traffic does not propagate "
            "across the network."
        ),
        "recommendation": (
            "It is recommended to apply an ACL to MLD snooping to filter unauthorized multicast groups.\n\n"
            "switch(config)# ipv6 mld snooping apply access list <ACL-NAME>"
        ),
        "expected_output": (
            "switch(config)# ipv6 mld snooping apply access list <ACL-NAME>"
        ),
        "check_fn": "check_mld_acl",
    },

    {
        "id": 49,
        "cis_id": "3.3.5",
        "obs_title": "MSDP Authentication & SA Filtering (Manual)",
        "observation": "It was observed that MSDP authentication and SA filtering are not configured.",
        "severity": "Medium",
        "description": (
            "The Multicast Source Discovery Protocol (MSDP) connects multiple PIM-SM domains, "
            "enabling RPs to share multicast source information using SA messages over TCP. "
            "MD5 authentication secures MSDP peer connections, while ACL-based SA filtering allows "
            "precise traffic control, improving network performance and mitigating security risks. "
            "MSDP is crucial for efficient interdomain multicast routing in multi-domain networks."
        ),
        "impact": (
            "MSDP enhances multicast scalability by enabling cross-domain source discovery while "
            "reducing configuration complexity. Without MD5 authentication, unauthorized devices "
            "could establish MSDP peer relationships. Without SA filtering, all multicast source "
            "announcements are shared, potentially overwhelming resources."
        ),
        "recommendation": (
            "It is recommended to configure MSDP peer authentication and SA filtering.\n\n"
            "switch(config)# router msdp\n"
            "switch(config-msdp)# ip msdp peer <ip-address>\n"
            "switch(config-msdp-peer)# password plaintext <password>\n"
            "switch(config-msdp-peer)# sa-filter in access-list <acl-name>\n"
            "switch(config-msdp-peer)# sa-filter out access-list <acl-name>"
        ),
        "expected_output": (
            "switch(config-msdp-peer)# password plaintext <password>\n"
            "switch(config-msdp-peer)# sa-filter in access-list <acl-name>"
        ),
        "check_fn": "check_msdp_auth",
    },

    {
        "id": 50,
        "cis_id": "3.3.6",
        "obs_title": "MSDP SA Cache limit (Manual)",
        "observation": "It was observed that MSDP SA cache limit is not configured.",
        "severity": "Low",
        "description": (
            "By default, the MSDP SA has no limit configured per peer. The sa-limit command limits "
            "the overall number of (S, G) entries that a device can accept from specified MSDP peers "
            "and store in a SA-cache. When configured, the device maintains a per-peer count of (S, G) "
            "messages stored in the SA-cache and ignores new messages from a peer if the configured "
            "sa-limit for that peer has been reached. This command protects MSDP-enabled devices from "
            "denial of service (DoS) attacks."
        ),
        "impact": (
            "Configuring the MSDP SA cache limit mitigates the risk of denial-of-service (DoS) "
            "attacks by restricting excessive multicast source entries. It prevents attackers from "
            "overwhelming the system with excessive SA messages, preserving network stability and "
            "resilience against DoS threats."
        ),
        "recommendation": (
            "It is recommended to configure an MSDP SA cache limit per peer to prevent DoS attacks.\n\n"
            "switch(config)# router msdp\n"
            "switch(config-msdp)# ip msdp peer <peer-address>\n"
            "switch(config-msdp)# sa-limit <value>"
        ),
        "expected_output": (
            "switch(config)# router msdp\n"
            "switch(config-msdp)# ip msdp peer <peer-address>\n"
            "switch(config-msdp)# sa-limit <value>"
        ),
        "check_fn": "check_msdp_sa_limit",
    },

    {
        "id": 51,
        "cis_id": "3.3.7",
        "obs_title": "Multicast Boundary ACL (Manual)",
        "observation": "It was observed that multicast boundary ACLs are not configured on the device.",
        "severity": "Low",
        "description": (
            "A multicast boundary ACL is a configuration mechanism that filters multicast traffic at "
            "specified interfaces to control the routing of multicast packets. This feature manages "
            "both multicast data and control packets including IGMP joins and PIM join/prune messages. "
            "It plays a crucial role in defining which multicast traffic is permitted or denied between "
            "different domains, preventing routing loops, managing traffic congestion, and ensuring "
            "necessary multicast traffic flows between domains."
        ),
        "impact": (
            "Configuring multicast boundary ACLs ensures that only authorized multicast traffic "
            "crosses domain boundaries, reducing the likelihood of network overload or disruptions. "
            "Without boundary ACLs, multicast traffic can propagate uncontrolled across the network, "
            "consuming resources and potentially disrupting operations in adjacent domains."
        ),
        "recommendation": (
            "It is recommended to configure multicast boundary ACLs on boundary interfaces.\n\n"
            "access-list ip bound\n"
            "    10 deny any any 239.0.0.0/255.0.0.0\n"
            "    20 permit any any 224.0.0.0/240.0.0.0\n"
            "switch(config)# interface <boundary-interface>\n"
            "switch(config-if)# ip multicast boundary access-list bound"
        ),
        "expected_output": (
            "switch(config)# interface <ID>\n"
            "switch(config-if)# ip multicast boundary access-list <ACL-NAME>"
        ),
        "check_fn": "check_multicast_boundary",
    },

    {
        "id": 52,
        "cis_id": "3.3.8",
        "obs_title": "Multicast BSR Boundary (Manual)",
        "observation": "It was observed that PIM BSR boundary is not configured on boundary interfaces.",
        "severity": "Low",
        "description": (
            "PIM Bootstrap messages (BSMs) play a crucial role in electing a Rendezvous Point (RP) "
            "for multicast routing within a domain. Exchanging these messages across different "
            "multicast domains is undesirable, as it can result in the unintended election of an RP "
            "in a domain where it does not belong, disrupting multicast operations. The bsr-boundary "
            "configuration on boundary interfaces ensures that PIM BSMs from external domains are "
            "dropped and internally generated BSMs are confined to their domain."
        ),
        "impact": (
            "Without BSR boundary configuration, PIM Bootstrap messages can cross domain boundaries "
            "and interfere with RP elections in other domains. This can lead to conflicts or "
            "misconfigurations where RPs are incorrectly elected in domains they are not supposed "
            "to serve, causing multicast routing inconsistencies and potential outages."
        ),
        "recommendation": (
            "It is recommended to configure PIM BSR boundary on all multicast domain boundary interfaces.\n\n"
            "switch(config)# interface <ID>\n"
            "switch(config-int)# ip pim-sparse bsr-boundary"
        ),
        "expected_output": (
            "switch(config)# interface <ID>\n"
            "switch(config-int)# ip pim-sparse bsr-boundary"
        ),
        "check_fn": "check_bsr_boundary",
    },

    # ── 3.4-3.8 Network Security ───────────────

    {
        "id": 53,
        "cis_id": "3.4",
        "obs_title": "IP Source Lockdown (Manual)",
        "observation": "It was observed that IP source lockdown is not configured on the device.",
        "severity": "Medium",
        "description": (
            "IP source lockdown provides added security by preventing IP source address spoofing on "
            "a per-port basis. Every packet is inspected in hardware. When IP source lockdown is "
            "enabled, IP traffic received on an interface is forwarded only if the VLAN, IP address, "
            "MAC address, and interface match the IP binding database entry. The binding database is "
            "dynamically populated by DHCP snooping or statically via the ip source-binding command."
        ),
        "impact": (
            "By implementing IP Source Lockdown, organizations can significantly reduce "
            "vulnerabilities to man-in-the-middle attacks, unauthorized device access, and network "
            "disruptions. This improves overall network integrity, enhances compliance with security "
            "policies, and ensures reliable and secure connectivity for critical business operations."
        ),
        "recommendation": (
            "It is recommended to enable IP source lockdown on untrusted interfaces.\n\n"
            "switch(config)# ip source-lockdown resource-extended\n"
            "switch(config)# interface <interface-id>\n"
            "switch(config-if)# ipv4 source-lockdown\n"
            "switch(config-if)# ipv6 source-lockdown"
        ),
        "expected_output": (
            "switch(config)# interface <interface-id>\n"
            "switch(config-if)# ipv4 source-lockdown"
        ),
        "check_fn": "check_ip_source_lockdown",
    },

    {
        "id": 54,
        "cis_id": "3.5",
        "obs_title": "Dynamic ARP Inspection (Manual)",
        "observation": "It was observed that Dynamic ARP Inspection is not configured on the device.",
        "severity": "Medium",
        "description": (
            "Dynamic ARP Inspection (DAI) is a security feature available in AOS-CX switches that "
            "validates ARP packets on the network to prevent ARP spoofing and ARP-based "
            "man-in-the-middle attacks. It verifies ARP requests and responses against trusted "
            "sources such as DHCP snooping and static ARP entries. DAI ensures that ARP traffic is "
            "authenticated before being forwarded, preventing attackers from injecting forged ARP "
            "packets to redirect traffic."
        ),
        "impact": (
            "Implementing Dynamic ARP Inspection enhances network security by blocking malicious "
            "ARP packets and mitigating risks of ARP-based attacks. It reduces downtime caused by "
            "compromised devices and ensures reliable communication in enterprise environments. "
            "Without DAI, attackers can perform ARP poisoning to intercept, modify, or disrupt "
            "network traffic."
        ),
        "recommendation": (
            "It is recommended to enable Dynamic ARP Inspection on all VLANs.\n\n"
            "switch(config)# vlan <vlan-id>\n"
            "switch(config-vlan)# arp inspection\n"
            "switch(config)# interface <interface-id>\n"
            "switch(config-if)# arp inspection trust"
        ),
        "expected_output": (
            "switch(config)# vlan <vlan-id>\n"
            "switch(config-vlan)# arp inspection\n"
            "switch(config-if)# arp inspection trust"
        ),
        "check_fn": "check_dai",
    },

    {
        "id": 55,
        "cis_id": "3.6",
        "obs_title": "ND Snooping (Manual)",
        "observation": "It was observed that ND Snooping is not configured on the device.",
        "severity": "Low",
        "description": (
            "ND (Neighbor Discovery) Snooping is a security feature that enhances IPv6 network "
            "integrity by monitoring and validating IPv6 neighbor discovery messages. It prevents "
            "malicious activities such as ND spoofing and rogue advertisements. ND snooping learns "
            "the source MAC addresses, source IPv6 addresses, input interfaces, and VLANs of "
            "incoming ND messages and data packets to build IP binding entries. ND snooping drops "
            "ND packets with mismatched MAC addresses or invalid IPv6 addresses."
        ),
        "impact": (
            "ND Snooping improves network security by mitigating IPv6-based attacks, enhances "
            "device discovery accuracy, and ensures reliable communication between endpoints in "
            "IPv6 environments. Without ND snooping, attackers can perform IPv6 neighbor discovery "
            "spoofing attacks, redirecting traffic or causing denial-of-service conditions."
        ),
        "recommendation": (
            "It is recommended to enable ND snooping globally and on all relevant VLANs.\n\n"
            "switch(config)# nd-snooping\n"
            "switch(config)# vlan <vlan-id>\n"
            "switch(config-vlan-id)# nd-snooping"
        ),
        "expected_output": (
            "switch(config)# nd-snooping\n"
            "switch(config-vlan-id)# nd-snooping"
        ),
        "check_fn": "check_nd_snooping",
    },

    {
        "id": 56,
        "cis_id": "3.7",
        "obs_title": "RA Guard (Manual)",
        "observation": "It was observed that RA Guard is not configured on the device.",
        "severity": "Low",
        "description": (
            "RA (Router Advertisement) Guard is a security feature in AOS-CX switches that monitors "
            "and filters IPv6 Router Advertisement messages, allowing only authorized RAs to "
            "propagate. It helps prevent rogue or malicious routers from disrupting network "
            "operations. When RA guard policy is enabled, RA packets received on trusted ports are "
            "validated against a set of parameters. RA Guard policy options include: Hop Limit, "
            "Managed Config Flag, Other Config Flag, Router Preference, ACL, and Advertised Prefix Lists."
        ),
        "impact": (
            "By deploying RA Guard, network administrators can prevent unauthorized devices from "
            "acting as routers, reducing the risk of network disruptions and enhancing overall IPv6 "
            "network stability, security, and performance. Without RA Guard, rogue router "
            "advertisements can redirect IPv6 traffic or disable IPv6 connectivity for legitimate hosts."
        ),
        "recommendation": (
            "It is recommended to enable RA Guard on all VLANs with untrusted ports.\n\n"
            "switch(config)# nd-snooping enable\n"
            "switch(config)# vlan <id>\n"
            "switch(config-vlan-id)# nd-snooping ra-guard"
        ),
        "expected_output": (
            "switch(config)# nd-snooping enable\n"
            "switch(config-vlan-id)# nd-snooping ra-guard"
        ),
        "check_fn": "check_ra_guard",
    },

    {
        "id": 57,
        "cis_id": "3.8",
        "obs_title": "IPv6 Destination Guard (Manual)",
        "observation": "It was observed that IPv6 Destination Guard is not configured on the device.",
        "severity": "Low",
        "description": (
            "The IPv6 Destination Guard feature in AOS-CX switches provides security by validating "
            "IPv6 traffic against a known database of allowed destinations. It ensures only authorized "
            "destination IP addresses are reachable, mitigating risks of unauthorized access or "
            "malicious activity within the network. This feature requires the binding table to be "
            "populated with the help of DHCPv6 snooping, ND snooping, or static-ip-bindings."
        ),
        "impact": (
            "Implementing IPv6 Destination Guard enhances network security by mitigating threats "
            "like IPv6 address spoofing and unauthorized data exfiltration. It contributes to a "
            "robust, reliable, and secure network environment by ensuring only legitimate IPv6 "
            "destinations are accessible. Without this feature, hosts can potentially reach "
            "unauthorized IPv6 destinations within the local network segment."
        ),
        "recommendation": (
            "It is recommended to enable IPv6 Destination Guard on all relevant VLANs.\n\n"
            "switch(config)# vlan <id>\n"
            "switch(config-vlan-<id>)# ipv6 destination-guard"
        ),
        "expected_output": (
            "switch(config)# vlan <id>\n"
            "switch(config-vlan-id)# ipv6 destination-guard"
        ),
        "check_fn": "check_ipv6_dest_guard",
    },

    # ══════════════════════════════════════════
    # SECTION 4 – CONTROL PLANE
    # ══════════════════════════════════════════

    # ── 4.1 COPP Policing ─────────────────────

    {
        "id": 58,
        "cis_id": "4.1.1",
        "obs_title": "Control Plane Policing (Manual)",
        "observation": "It was observed that a custom Control Plane Policing (CoPP) policy is not configured.",
        "severity": "Medium",
        "description": (
            "Control Plane Policing prevents flooding of certain types of packets from overloading "
            "the switch or module CPU by either rate-limiting or dropping packets. The switch software "
            "provides several configurable classes of packets that can be rate-limited, including ARP "
            "broadcasts, multicast, routing protocols (BGP, OSPF), and spanning tree. CoPP is always "
            "active and cannot be disabled. Proper CoPP configuration is required to protect against "
            "control plane denial-of-service attacks."
        ),
        "impact": (
            "CoPP is crucial for protecting the control plane and preventing network outages. "
            "Without proper CoPP configuration, the switch CPU can be overwhelmed by excessive "
            "control plane traffic, leading to performance degradation, network instability, and "
            "management access issues. This can be exploited by an attacker to cause a denial-of-"
            "service condition on the switch."
        ),
        "recommendation": (
            "It is recommended to configure and apply a custom CoPP policy appropriate for the environment.\n\n"
            "switch(config)# copp-policy <policy-name>\n"
            "switch(config-copp)# class <class> priority <0-6> rate <pps> burst <pkts>\n"
            "switch(config-copp)# exit\n"
            "switch(config)# apply copp-policy <policy-name>"
        ),
        "expected_output": (
            "switch(config)# copp-policy <policy-name>\n"
            "switch(config)# apply copp-policy <policy-name>"
        ),
        "check_fn": "check_copp_policy",
    },

    # ── 4.2 Spanning Tree Security ────────────

    {
        "id": 59,
        "cis_id": "4.2.1",
        "obs_title": "Spanning Tree BPDU Protect (Manual)",
        "observation": "It was observed that Spanning Tree BPDU Protect is not enabled on edge ports.",
        "severity": "Medium",
        "description": (
            "BPDU protection secures the active topology by preventing spoofed BPDU packets from "
            "entering the network. Typically, BPDU protection is applied on edge ports connected to "
            "end user devices that do not run STP. If STP BPDU packets are received on a protected "
            "port, BPDU guard disables the port and an alert is sent. Various security mechanisms "
            "are in place to protect spanning tree configurations from interference and rogue devices "
            "or unwarranted changes to the network."
        ),
        "impact": (
            "While BPDU protect will prevent spoofed BPDU traffic from entering the network, "
            "it must be correctly applied only on edge ports connected to end devices. If "
            "misconfigured on a port connected to another switch actively participating in STP, "
            "a loop can occur, leading to broadcast storms and network instability."
        ),
        "recommendation": (
            "It is recommended to enable BPDU Guard on all edge ports connected to end-user devices.\n\n"
            "switch(config)# interface <edge-port>\n"
            "switch(config-if)# no routing\n"
            "switch(config-if)# vlan access <vlan-id>\n"
            "switch(config-if)# spanning-tree bpdu-guard"
        ),
        "expected_output": (
            "switch(config)# interface <edge-port>\n"
            "switch(config-if)# spanning-tree bpdu-guard"
        ),
        "check_fn": "check_bpdu_guard",
    },

    {
        "id": 60,
        "cis_id": "4.2.2",
        "obs_title": "Spanning Tree Root Protect (Manual)",
        "observation": "It was observed that Spanning Tree Root Protect is not enabled on uplink ports.",
        "severity": "Medium",
        "description": (
            "Root protection secures the active topology by preventing other switches from declaring "
            "their ability to propagate superior BPDUs, which would normally replace the current root "
            "bridge selection. This is typically carried out between the core (required to be the root) "
            "and access switches to prevent ports that are not expected to originate root information "
            "(such as server ports and access switch ports) from becoming the root bridge."
        ),
        "impact": (
            "Enabling Root Guard prevents unauthorized switches from becoming the root bridge in "
            "your network's STP topology. Without Root Guard, a rogue or misconfigured switch could "
            "advertise superior BPDUs and become the root bridge, causing suboptimal traffic paths "
            "or network instability. If a port receives a superior BPDU when Root Guard is enabled, "
            "the port enters a root-inconsistent state and stops forwarding traffic."
        ),
        "recommendation": (
            "It is recommended to enable Spanning Tree Root Guard on uplink ports toward access switches.\n\n"
            "switch(config)# interface <uplink-port>\n"
            "switch(config-if)# vlan trunk allow all\n"
            "switch(config-if)# spanning-tree root-guard"
        ),
        "expected_output": (
            "switch(config)# interface <uplink-port>\n"
            "switch(config-if)# spanning-tree root-guard"
        ),
        "check_fn": "check_root_guard",
    },

    # ── 4.3 Control Plane ACL ─────────────────

    {
        "id": 61,
        "cis_id": "4.3.1",
        "obs_title": "Control Plane ACL Management (Manual)",
        "observation": "It was observed that no Control Plane ACL is applied to restrict management access.",
        "severity": "High",
        "description": (
            "Once an IP address is bound to an interface associated with a VRF, the switch may "
            "become exposed to management access from untrusted users or devices. This potential "
            "point of vulnerability can be mitigated by binding an Access Control List (ACL) to the "
            "control plane for that VRF. Once a control plane ACL is applied to a VRF, it filters "
            "packets to all IPv4/IPv6 addresses bound to the device on that VRF, restricting "
            "management access to trusted management devices only."
        ),
        "impact": (
            "There are significant risks associated with allowing any device with reachability to "
            "access the exposed ports on network infrastructure. Without a Control Plane ACL, any "
            "IP-reachable host can attempt to connect to management services such as SSH, HTTPS, "
            "and SNMP. Whitelisting specific trusted management devices reduces exposure to "
            "credential stuffing and man-in-the-middle (MITM) attacks."
        ),
        "recommendation": (
            "It is recommended to apply a Control Plane ACL to restrict management access to "
            "trusted hosts only.\n\n"
            "switch(config)# access-list ip <acl-name>\n"
            "switch(config-acl-ip)# permit ip <trusted-mgmt-network> any\n"
            "switch(config-acl-ip)# deny ip any any log\n"
            "switch(config)# apply access-list ip <acl-name> control-plane vrf mgmt"
        ),
        "expected_output": (
            "switch(config)# access-list ip <acl-name>\n"
            "switch(config-acl-ip)# permit ip <trusted-mgmt-network> any\n"
            "switch(config-acl-ip)# deny ip any any log\n"
            "switch(config)# apply access-list ip <acl-name> control-plane vrf mgmt"
        ),
        "check_fn": "check_cp_acl_mgmt",
    },
]


# ─────────────────────────────────────────────
# AUDIT CHECK FUNCTIONS
# Each returns: (passed: bool, found_lines: list of (line_no, line_text))
# ─────────────────────────────────────────────

def _exact_scan(lines, exact_patterns):
    """
    Scan every line against EXACT command patterns derived directly from the
    recommendation/expected-output for that check.
    Only lines that are a genuine match for those specific commands are returned.
    This ensures the Received Output POC contains only the relevant config lines
    and no extraneous text.
    """
    matched = {}
    for i, line in enumerate(lines, 1):
        for pat in exact_patterns:
            if re.search(pat, line, re.IGNORECASE):
                matched[i] = line.rstrip()
                break
    return sorted(matched.items())


# ── 1.1.3 Password Complexity ─────────────────────────────────────────────────
_PWD_INTENT = [r"\bpassword\s+complexity\b", r"\bminimum-length\b", r"\blowercase-count\b",
               r"\buppercase-count\b", r"\bspecial-char-count\b", r"\bnumeric-count\b", r"\bhistory-count\b"]
_PWD_PASS = [r"^\s*password\s+complexity", r"^\s*minimum-length\s+\d+", r"^\s*lowercase-count\s+\d+",
             r"^\s*uppercase-count\s+\d+", r"^\s*special-char-count\s+\d+",
             r"^\s*numeric-count\s+\d+", r"^\s*history-count\s+\d+"]


def check_password_complexity(lines):
    hits = {pat: False for pat in _PWD_PASS}
    for line in lines:
        for pat in _PWD_PASS:
            if not hits[pat] and re.search(pat, line, re.IGNORECASE):
                hits[pat] = True
    if all(hits.values()):
        return True, []
    return False, _exact_scan(lines, _PWD_PASS)


# ── 1.1.4 Export Password ─────────────────────────────────────────────────────
def check_export_password(lines):
    for line in lines:
        if re.search(r"export\s+password.*custom", line, re.IGNORECASE):
            return True, []
    return False, _exact_scan(lines, [r"^\s*service\s+export-password\b"])


# ── 1.1.8 Session Management ──────────────────────────────────────────────────
_SESSION_INTENT = [r"\bcli-session\b", r"\bmax-per-user\b", r"\btimeout\b", r"\btracking-range\b"]
_SESSION_PASS = [r"^\s*max-per-user\s+\d+", r"^\s*timeout\s+\d+", r"^\s*tracking-range\s+\d+"]


def check_session_management(lines):
    hits = {pat: False for pat in _SESSION_PASS}
    in_cli_session = False
    for line in lines:
        if re.search(r"^\s*cli-session", line, re.IGNORECASE):
            in_cli_session = True
        if in_cli_session:
            for pat in _SESSION_PASS:
                if not hits[pat] and re.search(pat, line, re.IGNORECASE):
                    hits[pat] = True
    if all(hits.values()):
        return True, []
    return False, _exact_scan(lines, _SESSION_PASS)


# ── 1.1.9 Telnet Disabled ─────────────────────────────────────────────────────
def check_telnet_disabled(lines):
    telnet_enabled = any(re.search(r"^\s*telnet\s+server\s+vrf\b", l, re.IGNORECASE) and
                         not re.search(r"^\s*no\s+telnet", l, re.IGNORECASE) for l in lines)
    no_telnet = any(re.search(r"^\s*no\s+telnet\s+server", l, re.IGNORECASE) for l in lines)
    if no_telnet or not telnet_enabled:
        return True, []
    return False, _exact_scan(lines, [r"^\s*telnet\s+server\s+vrf\b"])


# ── 1.2.1 SSH Public Key ──────────────────────────────────────────────────────
def check_ssh_pubkey(lines):
    if any(re.search(r"\bauthorized-key\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*user\s+\S+\s+authorized-key\b"])


# ── 1.2.2 SSH Allow List ──────────────────────────────────────────────────────
_SSH_INTENT = [r"\bssh\s+server\s+allow-list\b", r"\bssh\s+server\b"]
_SSH_PASS_BLOCK = r"ssh\s+server\s+allow-list"
_SSH_PASS_IP = r"^\s*ip\s+\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}"
_SSH_PASS_ENABLE = r"^\s*enable\s*$"


def check_ssh_allowlist(lines):
    has_block = any(re.search(_SSH_PASS_BLOCK, l, re.IGNORECASE) for l in lines)
    has_ip = any(re.search(_SSH_PASS_IP, l, re.IGNORECASE) for l in lines)
    has_enable = any(re.search(_SSH_PASS_ENABLE, l, re.IGNORECASE) for l in lines)
    if has_block and has_ip and has_enable:
        return True, []
    return False, _exact_scan(lines, [r"^\s*ssh\s+server\s+allow-list\b", r"^\s*ip\s+\d", r"^\s*enable\s*$"])


# ── 1.2.3 SSH Port ────────────────────────────────────────────────────────────
def check_ssh_port(lines):
    for line in lines:
        m = re.search(r"\bssh\s+server\s+port\s+(\d+)", line, re.IGNORECASE)
        if m and m.group(1) != "22":
            return True, []
    return False, _exact_scan(lines, [r"^\s*ssh\s+server\s+port\s+\d+"])


# ── 1.2.5 SSH 2FA ─────────────────────────────────────────────────────────────
def check_ssh_2fa(lines):
    has_cert = any(re.search(r"\bssh\s+certificate-as-authorized-key\b", l, re.IGNORECASE) for l in lines)
    has_2fa = any(re.search(r"\bssh\s+two-factor-authentication\b", l, re.IGNORECASE) for l in lines)
    if has_cert and has_2fa:
        return True, []
    return False, _exact_scan(lines, [r"^\s*ssh\s+certificate-as-authorized-key\b", r"^\s*ssh\s+two-factor-authentication\b"])


# ── 1.3.1 NTP Authentication ──────────────────────────────────────────────────
_NTP_INTENT = [r"\bntp\s+authentication\b", r"\bntp\s+authentication-key\b"]
_NTP_PASS_AUTH = r"^\s*ntp\s+authentication\s*$"
_NTP_PASS_KEY = r"^\s*ntp\s+authentication-key\s+"


def check_ntp_auth(lines):
    has_auth = any(re.search(_NTP_PASS_AUTH, l, re.IGNORECASE) for l in lines)
    has_key = any(re.search(_NTP_PASS_KEY, l, re.IGNORECASE) for l in lines)
    if has_auth and has_key:
        return True, []
    return False, _exact_scan(lines, [_NTP_PASS_AUTH, _NTP_PASS_KEY])


# ── 1.3.2 NTP Configured ─────────────────────────────────────────────────────
def check_ntp_configured(lines):
    has_server = any(re.search(r"^\s*ntp\s+server\s+\S+", l, re.IGNORECASE) for l in lines)
    has_enable = any(re.search(r"^\s*ntp\s+enable\s*$", l, re.IGNORECASE) for l in lines)
    if has_server and has_enable:
        return True, []
    return False, _exact_scan(lines, [r"^\s*ntp\s+server\s+\S+", r"^\s*ntp\s+enable\s*$", r"^\s*ntp\s+vrf\s+\S+"])


# ── 1.4.1.1 SNMP Community ────────────────────────────────────────────────────
def check_snmp_community(lines):
    has_non_default = False
    has_acl = False
    for line in lines:
        if re.search(r"\bsnmp-server\s+community\b", line, re.IGNORECASE):
            if not re.search(r"\bpublic\b|\bprivate\b", line, re.IGNORECASE):
                has_non_default = True
        if re.search(r"\baccess-list\b", line, re.IGNORECASE) and has_non_default:
            has_acl = True
    if has_non_default and has_acl:
        return True, []
    return False, _exact_scan(lines, [r"^\s*snmp-server\s+community\b", r"^\s*access-level\b", r"^\s*access-list\b"])


# ── 1.4.2.1 SNMPv3 ───────────────────────────────────────────────────────────
def check_snmpv3(lines):
    has_v3_user = any(re.search(r"\bsnmpv3\s+user\b", l, re.IGNORECASE) for l in lines)
    has_v3_only = any(re.search(r"\bsnmp-server\s+snmpv3-only\b", l, re.IGNORECASE) for l in lines)
    if has_v3_user and has_v3_only:
        return True, []
    return False, _exact_scan(lines, [r"^\s*snmpv3\s+user\b", r"^\s*snmp-server\s+snmpv3-only\b"])


# ── 1.4.3 SNMP Traps ─────────────────────────────────────────────────────────
def check_snmp_traps(lines):
    has_trap_config = any(re.search(r"\bsnmp-server\s+trap\s+configuration-changes\b", l, re.IGNORECASE) for l in lines)
    has_trap_host = any(re.search(r"\bsnmp-server\s+host\b.*\btrap\b", l, re.IGNORECASE) for l in lines)
    if has_trap_config and has_trap_host:
        return True, []
    return False, _exact_scan(lines, [r"^\s*snmp-server\s+trap\b", r"^\s*snmp-server\s+host\b"])


# ── 1.5.1.1 RADIUS Server ────────────────────────────────────────────────────
def check_radius_server(lines):
    has_radius = any(re.search(r"^\s*radius-server\s+host\b", l, re.IGNORECASE) for l in lines)
    has_group = any(re.search(r"^\s*aaa\s+group\s+server\s+radius\b", l, re.IGNORECASE) for l in lines)
    if has_radius and has_group:
        return True, []
    return False, _exact_scan(lines, [r"^\s*radius-server\s+host\b", r"^\s*aaa\s+group\s+server\s+radius\b"])


# ── 1.5.1.2 TACACS Server ────────────────────────────────────────────────────
def check_tacacs_server(lines):
    has_tacacs = any(re.search(r"^\s*tacacs-server\s+host\b", l, re.IGNORECASE) for l in lines)
    has_group = any(re.search(r"^\s*aaa\s+group\s+server\s+tacacs\b", l, re.IGNORECASE) for l in lines)
    if has_tacacs and has_group:
        return True, []
    return False, _exact_scan(lines, [r"^\s*tacacs-server\s+host\b", r"^\s*aaa\s+group\s+server\s+tacacs\b"])


# ── 1.5.2.2 Limit Login Attempts ─────────────────────────────────────────────
_AAA_PASS_CONSOLE = r"aaa\s+authentication\s+console-login-attempts\s+\d+\s+console-lockout-time\s+\d+"
_AAA_PASS_LIMIT = r"aaa\s+authentication\s+limit-login-attempts\s+\d+\s+lockout-time\s+\d+"


def check_failed_attempts(lines):
    has_console = any(re.search(_AAA_PASS_CONSOLE, l, re.IGNORECASE) for l in lines)
    has_limit = any(re.search(_AAA_PASS_LIMIT, l, re.IGNORECASE) for l in lines)
    if has_console and has_limit:
        return True, []
    return False, _exact_scan(lines, [_AAA_PASS_CONSOLE, _AAA_PASS_LIMIT])


# ── 1.5.2.3 Remote Authentication ────────────────────────────────────────────
def check_remote_auth(lines):
    if any(re.search(r"^\s*aaa\s+authentication\s+login\s+(default|ssh|console|telnet|https-server)\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*aaa\s+authentication\s+login\s+(default|ssh|console|telnet|https-server)\b"])


# ── 1.5.3.1 Local Authorization ──────────────────────────────────────────────
def check_local_authorization(lines):
    if any(re.search(r"^\s*aaa\s+authorization\s+commands\b.*\blocal\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*aaa\s+authorization\s+commands\b.*\blocal\b"])


# ── 1.5.3.2 Remote Authorization ─────────────────────────────────────────────
def check_remote_authorization(lines):
    for line in lines:
        if re.search(r"^\s*aaa\s+authorization\s+commands\b", line, re.IGNORECASE):
            if re.search(r"\bgroup\b", line, re.IGNORECASE):
                return True, []
    return False, _exact_scan(lines, [r"^\s*aaa\s+authorization\s+commands\b.*\bgroup\b"])


# ── 1.5.4.1 Local Accounting ─────────────────────────────────────────────────
def check_local_accounting(lines):
    if any(re.search(r"^\s*aaa\s+accounting\s+commands\b.*\blocal\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*aaa\s+accounting\s+commands\b.*\blocal\b"])


# ── 1.5.4.2 Remote Accounting ────────────────────────────────────────────────
def check_remote_accounting(lines):
    for line in lines:
        if re.search(r"^\s*aaa\s+accounting\s+all-mgmt\b", line, re.IGNORECASE):
            if re.search(r"\bstart-stop\b", line, re.IGNORECASE):
                return True, []
    return False, _exact_scan(lines, [r"^\s*aaa\s+accounting\s+all-mgmt\b"])


# ── 1.5.6 Privilege Elevation ────────────────────────────────────────────────
def check_privilege_elevation(lines):
    if any(re.search(r"\baaa\s+authentication\s+login\s+privilege-elevation\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*aaa\s+authentication\s+login\s+privilege-elevation\b"])


# ── 1.6.2 TLS Version ────────────────────────────────────────────────────────
def check_tls_version(lines):
    if any(re.search(r"\btls\s+minimum-version\s+tls1[23]\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*tls\s+minimum-version\b"])


# ── 1.9.1 HTTPS Server ───────────────────────────────────────────────────────
def check_https_server(lines):
    disabled = any(re.search(r"^\s*no\s+https-server\s+vrf\b", l, re.IGNORECASE) for l in lines)
    if disabled:
        return True, []
    return False, _exact_scan(lines, [r"^\s*https-server\s+vrf\b", r"^\s*no\s+https-server\s+vrf\b"])


# ── 1.9.2 HTTPS Timeout ──────────────────────────────────────────────────────
def check_https_timeout(lines):
    for line in lines:
        m = re.search(r"\bhttps-server\s+session-timeout\s+(\d+)", line, re.IGNORECASE)
        if m and int(m.group(1)) <= 15:
            return True, []
    return False, _exact_scan(lines, [r"^\s*https-server\s+session-timeout\s+\d+"])


# ── 1.10.1 ServiceOS Password ────────────────────────────────────────────────
def check_serviceos_password(lines):
    if any(re.search(r"^\s*system\s+serviceos\s+password-prompt\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*system\s+serviceos\s+password-prompt\b"])


# ── 1.11.2 Syslog TLS ────────────────────────────────────────────────────────
_SYSLOG_PASS = r"^\s*logging\s+\S+\s+tls\s+\S+\s+auth-mode\s+subject-name\s+vrf\s+mgmt\s+include-auditable-events"
_SYSLOG_INTENT = [r"\blogging\s+\S+\s+tls\s+\S+\s+auth-mode\s+subject-name\s+vrf\s+mgmt\s+include-auditable-events\b"]


def check_syslog(lines):
    if any(re.search(_SYSLOG_PASS, l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [_SYSLOG_PASS])


# ── 1.12 Login Banner ────────────────────────────────────────────────────────
def check_login_banner(lines):
    has_motd = any(re.search(r"^\s*banner\s+motd\b", l, re.IGNORECASE) for l in lines)
    if has_motd:
        return True, []
    return False, _exact_scan(lines, [r"^\s*banner\s+motd\b", r"^\s*banner\s+exec\b"])


# ── 1.13 Config Backup ────────────────────────────────────────────────────────
def check_config_backup(lines):
    has_job = any(re.search(r"^\s*job\s+\S+", l, re.IGNORECASE) for l in lines)
    has_schedule = any(re.search(r"^\s*schedule\s+\S+", l, re.IGNORECASE) for l in lines)
    if has_job and has_schedule:
        return True, []
    return False, _exact_scan(lines, [r"^\s*job\s+\S+", r"^\s*schedule\s+\S+"])


# ── 1.14 Hostname ─────────────────────────────────────────────────────────────
def check_hostname(lines):
    for line in lines:
        m = re.search(r"^\s*hostname\s+(\S+)", line, re.IGNORECASE)
        if m and m.group(1).lower() not in ("switch", "localhost", "router", "aruba"):
            return True, []
    return False, _exact_scan(lines, [r"^\s*hostname\s+\S+"])


# ── 2.1.1 USB/Bluetooth ───────────────────────────────────────────────────────
def check_usb_bluetooth(lines):
    no_usb = any(re.search(r"^\s*no\s+usb\b", l, re.IGNORECASE) for l in lines)
    bt_disabled = any(re.search(r"^\s*bluetooth\s+disable\b", l, re.IGNORECASE) for l in lines)
    if no_usb and bt_disabled:
        return True, []
    return False, _exact_scan(lines, [r"^\s*no\s+usb\b", r"^\s*bluetooth\s+disable\b"])


# ── 2.1.3 Unused Interfaces ──────────────────────────────────────────────────
def check_unused_interfaces(lines):
    has_shutdown = any(re.search(r"^\s*shutdown\b", l, re.IGNORECASE) for l in lines)
    if has_shutdown:
        return True, []
    return False, _exact_scan(lines, [r"^\s*shutdown\b"])


# ── 2.2 Rate Limiting ─────────────────────────────────────────────────────────
def check_rate_limiting(lines):
    if any(re.search(r"^\s*rate-limit\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*rate-limit\s+(broadcast|multicast|unknown-unicast|icmp)\b"])


# ── 2.3 Proxy ARP ─────────────────────────────────────────────────────────────
def check_proxy_arp(lines):
    enabled = any(re.search(r"^\s*(ip\s+proxy-arp|ip\s+local-proxy-arp|ipv6\s+local-proxy-nd)\b", l, re.IGNORECASE)
                  and not re.search(r"^\s*no\s+", l) for l in lines)
    if not enabled:
        return True, []
    return False, _exact_scan(lines, [r"^\s*ip\s+proxy-arp\b", r"^\s*ip\s+local-proxy-arp\b", r"^\s*ipv6\s+local-proxy-nd\b"])


# ── 2.4 IP Directed Broadcast ─────────────────────────────────────────────────
def check_directed_broadcast(lines):
    enabled = any(re.search(r"^\s*ip\s+directed-broadcast\b", l, re.IGNORECASE)
                  and not re.search(r"^\s*no\s+", l) for l in lines)
    if not enabled:
        return True, []
    return False, _exact_scan(lines, [r"^\s*ip\s+directed-broadcast\b"])


# ── 3.1.1.1 OSPF Passive ─────────────────────────────────────────────────────
def check_ospf_passive(lines):
    if any(re.search(r"^\s*passive-interface\s+default\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*passive-interface\s+default\b"])


# ── 3.1.1.2 OSPF Auth ─────────────────────────────────────────────────────────
def check_ospf_auth(lines):
    has_keychain = any(re.search(r"^\s*keychain\s+\S+", l, re.IGNORECASE) for l in lines)
    has_ospf_auth = any(re.search(r"\bip\s+ospf\s+authentication\s+keychain\b", l, re.IGNORECASE) for l in lines)
    if has_keychain and has_ospf_auth:
        return True, []
    return False, _exact_scan(lines, [r"^\s*keychain\s+\S+", r"^\s*cryptographic-algorithm\b", r"^\s*ip\s+ospf\s+authentication\s+keychain\b", r"^\s*ip\s+ospf\s+keychain\b"])


# ── 3.1.1.3 OSPFv3 IPsec ─────────────────────────────────────────────────────
def check_ospfv3_ipsec(lines):
    if any(re.search(r"\barea\s+\S+\s+(authentication|encryption)\s+ipsec\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*area\s+\S+\s+(authentication|encryption)\s+ipsec\b"])


# ── 3.1.2.1 BGP CP ACL ───────────────────────────────────────────────────────
def check_bgp_cp_acl(lines):
    has_acl_bgp = any(re.search(r"\beq\s+179\b", l, re.IGNORECASE) for l in lines)
    has_cp_apply = any(re.search(r"\bapply\s+access-list\b.*\bcontrol-plane\b", l, re.IGNORECASE) for l in lines)
    if has_acl_bgp and has_cp_apply:
        return True, []
    return False, _exact_scan(lines, [r"\beq\s+179\b", r"^\s*apply\s+access-list\b.*\bcontrol-plane\b"])


# ── 3.1.2.2 BGP MD5 ──────────────────────────────────────────────────────────
def check_bgp_md5(lines):
    if any(re.search(r"\bneighbor\b.*\bpassword\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*neighbor\s+\S+\s+password\b"])


# ── 3.1.2.3 BGP TTL Security ─────────────────────────────────────────────────
def check_bgp_ttl(lines):
    if any(re.search(r"\bneighbor\b.*\bttl-security-hops\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*neighbor\s+\S+\s+ttl-security-hops\b"])


# ── 3.2.1.1 DHCP Snooping ────────────────────────────────────────────────────
def check_dhcp_snooping(lines):
    has_global = any(re.search(r"^\s*dhcpv4-snooping\s*$", l, re.IGNORECASE) for l in lines)
    has_trust = any(re.search(r"^\s*dhcpv4-snooping\s+trust\b", l, re.IGNORECASE) for l in lines)
    if has_global and has_trust:
        return True, []
    return False, _exact_scan(lines, [r"^\s*dhcpv4-snooping\b", r"^\s*dhcpv6-snooping\b"])


# ── 3.2.1.2 DHCPv6 Guard ─────────────────────────────────────────────────────
def check_dhcpv6_guard(lines):
    if any(re.search(r"\bdhcpv6-snooping\s+guard-policy\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*dhcpv6-snooping\s+guard-policy\b"])


# ── 3.3.1.1 PIM Accept-Register ──────────────────────────────────────────────
def check_pim_accept_register(lines):
    if any(re.search(r"\baccept-register\s+access-list\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*accept-register\s+access-list\b"])


# ── 3.3.1.2 PIM Accept-RP ────────────────────────────────────────────────────
def check_pim_accept_rp(lines):
    if any(re.search(r"\baccept-rp\b.*\baccess-list\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*accept-rp\s+\S+\s+access-list\b"])


# ── 3.3.1.3 PIM SSM ──────────────────────────────────────────────────────────
def check_pim_ssm(lines):
    if any(re.search(r"\bpim-ssm\s+range-access-list\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*pim-ssm\s+range-access-list\b"])


# ── 3.3.3 IGMP ACL ───────────────────────────────────────────────────────────
def check_igmp_acl(lines):
    if any(re.search(r"\bip\s+igmp\s+snooping\s+apply\s+access\s+list\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*ip\s+igmp\s+snooping\s+apply\s+access\s+list\b"])


# ── 3.3.4 MLD ACL ────────────────────────────────────────────────────────────
def check_mld_acl(lines):
    if any(re.search(r"\bipv6\s+mld\s+snooping\s+apply\s+access\s+list\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*ipv6\s+mld\s+snooping\s+apply\s+access\s+list\b"])


# ── 3.3.5 MSDP Auth ──────────────────────────────────────────────────────────
def check_msdp_auth(lines):
    has_peer = any(re.search(r"\bip\s+msdp\s+peer\b", l, re.IGNORECASE) for l in lines)
    has_pass = any(re.search(r"\bpassword\s+plaintext\b", l, re.IGNORECASE) for l in lines)
    has_filter = any(re.search(r"\bsa-filter\b", l, re.IGNORECASE) for l in lines)
    if has_peer and has_pass and has_filter:
        return True, []
    return False, _exact_scan(lines, [r"^\s*ip\s+msdp\s+peer\b", r"^\s*password\s+plaintext\b", r"^\s*sa-filter\b"])


# ── 3.3.6 MSDP SA Limit ──────────────────────────────────────────────────────
def check_msdp_sa_limit(lines):
    if any(re.search(r"^\s*sa-limit\s+\d+", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*sa-limit\s+\d+"])


# ── 3.3.7 Multicast Boundary ─────────────────────────────────────────────────
def check_multicast_boundary(lines):
    if any(re.search(r"\bip\s+multicast\s+boundary\s+access-list\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*ip\s+multicast\s+boundary\s+access-list\b"])


# ── 3.3.8 BSR Boundary ────────────────────────────────────────────────────────
def check_bsr_boundary(lines):
    if any(re.search(r"\bip\s+pim-sparse\s+bsr-boundary\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*ip\s+pim-sparse\s+bsr-boundary\b"])


# ── 3.4 IP Source Lockdown ───────────────────────────────────────────────────
def check_ip_source_lockdown(lines):
    if any(re.search(r"\bipv4\s+source-lockdown\b|\bipv6\s+source-lockdown\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*ipv4\s+source-lockdown\b", r"^\s*ipv6\s+source-lockdown\b"])


# ── 3.5 DAI ──────────────────────────────────────────────────────────────────
def check_dai(lines):
    has_arp_inspection = any(re.search(r"^\s*arp\s+inspection\s*$", l, re.IGNORECASE) for l in lines)
    has_trust = any(re.search(r"^\s*arp\s+inspection\s+trust\b", l, re.IGNORECASE) for l in lines)
    if has_arp_inspection and has_trust:
        return True, []
    return False, _exact_scan(lines, [r"^\s*arp\s+inspection\b"])


# ── 3.6 ND Snooping ──────────────────────────────────────────────────────────
def check_nd_snooping(lines):
    has_global = any(re.search(r"^\s*nd-snooping\s*$", l, re.IGNORECASE) for l in lines)
    if has_global:
        return True, []
    return False, _exact_scan(lines, [r"^\s*nd-snooping\b"])


# ── 3.7 RA Guard ─────────────────────────────────────────────────────────────
def check_ra_guard(lines):
    if any(re.search(r"\bnd-snooping\s+ra-guard\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*nd-snooping\s+ra-guard\b", r"^\s*nd-snooping\s+enable\b"])


# ── 3.8 IPv6 Destination Guard ───────────────────────────────────────────────
def check_ipv6_dest_guard(lines):
    if any(re.search(r"\bipv6\s+destination-guard\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*ipv6\s+destination-guard\b"])


# ── 4.1.1 CoPP ───────────────────────────────────────────────────────────────
def check_copp_policy(lines):
    has_custom = any(re.search(r"^\s*copp-policy\s+(?!default)\S+", l, re.IGNORECASE) for l in lines)
    has_apply = any(re.search(r"^\s*apply\s+copp-policy\b", l, re.IGNORECASE) for l in lines)
    if has_custom and has_apply:
        return True, []
    return False, _exact_scan(lines, [r"^\s*copp-policy\s+\S+", r"^\s*apply\s+copp-policy\b"])


# ── 4.2.1 BPDU Guard ─────────────────────────────────────────────────────────
def check_bpdu_guard(lines):
    if any(re.search(r"\bspanning-tree\s+bpdu-guard\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*spanning-tree\s+bpdu-guard\b"])


# ── 4.2.2 Root Guard ─────────────────────────────────────────────────────────
def check_root_guard(lines):
    if any(re.search(r"\bspanning-tree\s+root-guard\b", l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _exact_scan(lines, [r"^\s*spanning-tree\s+root-guard\b"])


# ── 4.3.1 Control Plane ACL Mgmt ─────────────────────────────────────────────
def check_cp_acl_mgmt(lines):
    has_acl = any(re.search(r"^\s*access-list\s+ip\s+\S+", l, re.IGNORECASE) for l in lines)
    has_apply = any(re.search(r"\bapply\s+access-list\b.*\bcontrol-plane\b.*\bvrf\b", l, re.IGNORECASE) for l in lines)
    if has_acl and has_apply:
        return True, []
    return False, _exact_scan(lines, [r"^\s*access-list\s+ip\s+\S+", r"^\s*apply\s+access-list\b.*\bcontrol-plane\b"])


CHECK_DISPATCH = {
    "check_password_complexity":  check_password_complexity,
    "check_export_password":      check_export_password,
    "check_session_management":   check_session_management,
    "check_telnet_disabled":      check_telnet_disabled,
    "check_ssh_pubkey":           check_ssh_pubkey,
    "check_ssh_allowlist":        check_ssh_allowlist,
    "check_ssh_port":             check_ssh_port,
    "check_ssh_2fa":              check_ssh_2fa,
    "check_ntp_auth":             check_ntp_auth,
    "check_ntp_configured":       check_ntp_configured,
    "check_snmp_community":       check_snmp_community,
    "check_snmpv3":               check_snmpv3,
    "check_snmp_traps":           check_snmp_traps,
    "check_radius_server":        check_radius_server,
    "check_tacacs_server":        check_tacacs_server,
    "check_failed_attempts":      check_failed_attempts,
    "check_remote_auth":          check_remote_auth,
    "check_local_authorization":  check_local_authorization,
    "check_remote_authorization": check_remote_authorization,
    "check_local_accounting":     check_local_accounting,
    "check_remote_accounting":    check_remote_accounting,
    "check_privilege_elevation":  check_privilege_elevation,
    "check_tls_version":          check_tls_version,
    "check_https_server":         check_https_server,
    "check_https_timeout":        check_https_timeout,
    "check_serviceos_password":   check_serviceos_password,
    "check_syslog":               check_syslog,
    "check_login_banner":         check_login_banner,
    "check_config_backup":        check_config_backup,
    "check_hostname":             check_hostname,
    "check_usb_bluetooth":        check_usb_bluetooth,
    "check_unused_interfaces":    check_unused_interfaces,
    "check_rate_limiting":        check_rate_limiting,
    "check_proxy_arp":            check_proxy_arp,
    "check_directed_broadcast":   check_directed_broadcast,
    "check_ospf_passive":         check_ospf_passive,
    "check_ospf_auth":            check_ospf_auth,
    "check_ospfv3_ipsec":         check_ospfv3_ipsec,
    "check_bgp_cp_acl":           check_bgp_cp_acl,
    "check_bgp_md5":              check_bgp_md5,
    "check_bgp_ttl":              check_bgp_ttl,
    "check_dhcp_snooping":        check_dhcp_snooping,
    "check_dhcpv6_guard":         check_dhcpv6_guard,
    "check_pim_accept_register":  check_pim_accept_register,
    "check_pim_accept_rp":        check_pim_accept_rp,
    "check_pim_ssm":              check_pim_ssm,
    "check_igmp_acl":             check_igmp_acl,
    "check_mld_acl":              check_mld_acl,
    "check_msdp_auth":            check_msdp_auth,
    "check_msdp_sa_limit":        check_msdp_sa_limit,
    "check_multicast_boundary":   check_multicast_boundary,
    "check_bsr_boundary":         check_bsr_boundary,
    "check_ip_source_lockdown":   check_ip_source_lockdown,
    "check_dai":                  check_dai,
    "check_nd_snooping":          check_nd_snooping,
    "check_ra_guard":             check_ra_guard,
    "check_ipv6_dest_guard":      check_ipv6_dest_guard,
    "check_copp_policy":          check_copp_policy,
    "check_bpdu_guard":           check_bpdu_guard,
    "check_root_guard":           check_root_guard,
    "check_cp_acl_mgmt":          check_cp_acl_mgmt,
}


# ─────────────────────────────────────────────
# HELPER: extract device IP from filename
# ─────────────────────────────────────────────

def extract_device_ip(filepath, lines):
    basename = os.path.basename(filepath)
    ip_match = re.search(r"(\d{1,3}(?:\.\d{1,3}){3})", basename)
    if ip_match:
        return ip_match.group(1)
    for line in lines[:10]:
        ip_match = re.search(r"(\d{1,3}(?:\.\d{1,3}){3})", line)
        if ip_match:
            return ip_match.group(1)
    return "Unknown"


# ─────────────────────────────────────────────
# CORE AUDIT RUNNER
# ─────────────────────────────────────────────

def run_audit(filepath):
    with open(filepath, "r", errors="replace") as f:
        lines = f.readlines()

    device_ip = extract_device_ip(filepath, lines)
    findings = []

    for check in AUDIT_CHECKS:
        fn = CHECK_DISPATCH[check["check_fn"]]
        passed, found_lines = fn(lines)

        if passed:
            continue

        def _is_meaningful(raw):
            s = raw.strip()
            if not s:
                return False
            if s == "!":
                return False
            if re.match(r"^!\s*[─═\-]{2,}", s):
                return False
            return True

        def _display_text(raw):
            return raw.strip()

        meaningful = [(ln, lt) for ln, lt in found_lines if _is_meaningful(lt)]

        if not meaningful:
            received = (
                "The config line that needs to be checked:\n"
                "\n"
                "Missing Configuration"
            )
        else:
            config_lines = "\n".join(
                f"hostname(config)# {_display_text(lt)}" for _, lt in meaningful
            )
            line_nums = [ln for ln, _ in meaningful]
            ref = (
                f"Reference Line: {line_nums[0]}"
                if len(line_nums) == 1
                else f"Reference Line: {line_nums[0]} - {line_nums[-1]}"
            )
            received = (
                "The config line that needs to be checked:\n"
                "\n"
                f"{config_lines}\n"
                "\n"
                f"{ref}"
            )

        expected = (
            "The expected config line that needs to be altered:\n"
            "\n"
            "\n"
            f"{check['expected_output']}"
        )

        findings.append({
            "cis_id":      check["cis_id"],
            "severity":    check["severity"],
            "obs_title":   check["obs_title"],
            "observation": check["observation"],
            "device":      device_ip,
            "description": check["description"],
            "impact":      check["impact"],
            "recommendation": check["recommendation"],
            "received":    received,
            "expected":    expected,
        })

    return findings, device_ip


# ─────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FONT = Font(bold=True, color="000000", name="Arial", size=10)
CELL_FONT = Font(bold=False, name="Arial", size=9)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
_THIN = Side(style="thin", color="000000")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
NO_FILL = PatternFill(fill_type=None)
SEV_FILL = {
    "High":   PatternFill("solid", fgColor="FF0000"),
    "Medium": PatternFill("solid", fgColor="F5C242"),
    "Low":    PatternFill("solid", fgColor="00B050"),
}
CENTER_TOP_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)

COLUMNS = [
    ("#",               6),
    ("CIS ID",         10),
    ("Severity",       12),
    ("Observation",    35),
    ("Affected Device",18),
    ("Description",    40),
    ("Impact",         40),
    ("Recommendation", 45),
    ("Received Output",45),
    ("Expected Output",45),
    ("FSL Remarks",    20),
    ("EY Remarks",     20),
]


def _write_sheet(wb, findings, sheet_name):
    safe_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)[:31]
    existing = [s.title for s in wb.worksheets]
    if safe_name in existing:
        suffix = 2
        while f"{safe_name[:28]}_{suffix}" in existing:
            suffix += 1
        safe_name = f"{safe_name[:28]}_{suffix}"

    ws = wb.create_sheet(title=safe_name)
    ws.sheet_view.showGridLines = False

    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    if not findings:
        ws.cell(row=2, column=1,
                value="All checks PASSED — no issues found for this device.")
        ws.cell(row=2, column=1).font = Font(name="Arial", size=10, bold=True, color="107C10")
        ws.row_dimensions[2].height = 30
        return

    for row_idx, f in enumerate(findings, 2):
        obs_rich = CellRichText(
            TextBlock(InlineFont(b=True, rFont="Arial", sz=9), f["obs_title"]),
            TextBlock(InlineFont(b=False, rFont="Arial", sz=9), "\n\n\n" + f["observation"]),
        )

        values = [
            row_idx - 1,          # col A  #
            f["cis_id"],          # col B  CIS ID
            f["severity"],        # col C  Severity
            obs_rich,             # col D  Observation (rich text)
            f["device"],          # col E  Affected Device
            f["description"],     # col F  Description
            f["impact"],          # col G  Impact
            f["recommendation"],  # col H  Recommendation
            f["received"],        # col I  Received Output
            f["expected"],        # col J  Expected Output
            "",                   # col K  FSL Remarks
            "",                   # col L  EY Remarks
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER
            cell.fill = NO_FILL

            if col_idx == 3:                       # Severity
                cell.font = Font(bold=True, name="Arial", size=9)
                cell.fill = SEV_FILL.get(f["severity"], NO_FILL)
                cell.alignment = CENTER_TOP_ALIGN
            elif col_idx == 4:                     # Observation (rich text)
                cell.alignment = DATA_ALIGN
            elif col_idx in (1, 2, 5):             # #, CIS ID, Affected Device
                cell.font = CELL_FONT
                cell.alignment = CENTER_TOP_ALIGN
            else:
                cell.font = CELL_FONT
                cell.alignment = DATA_ALIGN

        ws.row_dimensions[row_idx].height = 180


def write_multi_excel(results, out_path):
    wb = Workbook()
    wb.remove(wb.active)
    for device_ip, findings in results:
        _write_sheet(wb, findings, device_ip)
    wb.save(out_path)


def write_excel(findings, out_path, device_ip):
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, findings, device_ip)
    wb.save(out_path)


# ─────────────────────────────────────────────
# TKINTER GUI
# ─────────────────────────────────────────────

class AuditApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Aruba OS – Secure Configuration Auditor  |  CIS Benchmark v1.0.1")
        self.resizable(True, True)
        self.minsize(620, 620)
        self.configure(bg="#1F3864")
        self._files = []
        self._build_ui()

    def _build_ui(self):
        title_frame = tk.Frame(self, bg="#1F3864", pady=14)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame,
            text="🔒  Aruba OS  |  Secure Configuration Auditor",
            bg="#1F3864", fg="white",
            font=("Arial", 16, "bold"),
        ).pack()
        tk.Label(
            title_frame,
            text="CIS HPE Aruba Networking CX Switch Benchmark v1.0.1  |  61 Checks",
            bg="#1F3864", fg="#A0B4D0",
            font=("Arial", 9),
        ).pack()

        card = tk.Frame(self, bg="white", padx=28, pady=24, relief="flat")
        card.pack(padx=24, pady=(0, 24), fill="both", expand=True)

        tk.Label(card, text="Configuration Log Files", bg="white",
                 font=("Arial", 10, "bold"), anchor="w").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 4))

        list_frame = tk.Frame(card, bg="white")
        list_frame.grid(row=1, column=0, columnspan=3, sticky="nsew", pady=(0, 4))

        self.file_list = tk.Listbox(
            list_frame, height=6, font=("Arial", 9),
            relief="solid", bd=1, bg="#F4F6F8",
            selectmode=tk.EXTENDED, activestyle="none"
        )
        list_sb = ttk.Scrollbar(list_frame, orient="vertical", command=self.file_list.yview)
        self.file_list.configure(yscrollcommand=list_sb.set)
        self.file_list.pack(side="left", fill="both", expand=True)
        list_sb.pack(side="right", fill="y")

        btn_frame = tk.Frame(card, bg="white")
        btn_frame.grid(row=2, column=0, columnspan=3, sticky="w", pady=(0, 6))

        tk.Button(btn_frame, text="＋  Add Files", command=self._add_files,
                  bg="#2E5BA8", fg="white", font=("Arial", 9, "bold"),
                  relief="flat", cursor="hand2", padx=10, pady=4).pack(side="left", padx=(0, 8))

        tk.Button(btn_frame, text="✕  Remove Selected", command=self._remove_selected,
                  bg="#C0392B", fg="white", font=("Arial", 9, "bold"),
                  relief="flat", cursor="hand2", padx=10, pady=4).pack(side="left", padx=(0, 8))

        tk.Button(btn_frame, text="⊘  Clear All", command=self._clear_files,
                  bg="#7F8C8D", fg="white", font=("Arial", 9, "bold"),
                  relief="flat", cursor="hand2", padx=10, pady=4).pack(side="left")

        self.file_count_var = tk.StringVar(value="0 file(s) selected")
        tk.Label(card, textvariable=self.file_count_var, bg="white",
                 fg="#888", font=("Arial", 8, "italic")).grid(
            row=3, column=0, columnspan=3, sticky="w")

        tk.Label(card, text="Output Folder", bg="white",
                 font=("Arial", 10, "bold"), anchor="w").grid(
            row=4, column=0, columnspan=3, sticky="w", pady=(14, 4))

        self.out_var = tk.StringVar(
            value=os.path.expanduser("~\\Desktop") if os.name == "nt"
            else os.path.expanduser("~/Desktop")
        )
        tk.Entry(card, textvariable=self.out_var, width=48,
                 font=("Arial", 9), relief="solid", bd=1).grid(
            row=5, column=0, columnspan=2, sticky="ew", ipady=4)

        tk.Button(card, text="  Browse…  ", command=self._browse_output,
                  bg="#2E5BA8", fg="white", font=("Arial", 9, "bold"),
                  relief="flat", cursor="hand2", padx=8, pady=4).grid(row=5, column=2, padx=(8, 0))

        tk.Label(card, text="Output Report Name", bg="white",
                 font=("Arial", 10, "bold"), anchor="w").grid(
            row=6, column=0, columnspan=3, sticky="w", pady=(12, 4))

        self.report_name_var = tk.StringVar(value="AuditReport_MultiDevice")
        tk.Entry(card, textvariable=self.report_name_var, width=48,
                 font=("Arial", 9), relief="solid", bd=1).grid(
            row=7, column=0, columnspan=2, sticky="ew", ipady=4)
        tk.Label(card, text=".xlsx", bg="white",
                 font=("Arial", 9), fg="#555").grid(row=7, column=2, sticky="w", padx=(6, 0))

        ttk.Separator(card, orient="horizontal").grid(
            row=8, column=0, columnspan=3, sticky="ew", pady=18)

        self.run_btn = tk.Button(
            card, text="▶   Run Audit", command=self._start_audit,
            bg="#107C10", fg="white", font=("Arial", 11, "bold"),
            relief="flat", cursor="hand2", padx=16, pady=8, width=20
        )
        self.run_btn.grid(row=9, column=0, columnspan=3)

        self.progress = ttk.Progressbar(card, mode="indeterminate", length=460)
        self.progress.grid(row=10, column=0, columnspan=3, pady=(14, 4))

        self.status_var = tk.StringVar(value="Ready. Add .log/.conf files to begin.")
        tk.Label(card, textvariable=self.status_var, bg="white",
                 fg="#555", font=("Arial", 9, "italic")).grid(
            row=11, column=0, columnspan=3, pady=(0, 6))

        tk.Label(card, text="Findings Preview", bg="white",
                 font=("Arial", 10, "bold"), anchor="w").grid(
            row=12, column=0, columnspan=3, sticky="w", pady=(10, 4))

        cols = ("#", "CIS ID", "Severity", "Observation", "Affected Device")
        self.tree = ttk.Treeview(card, columns=cols, show="headings",
                                 height=7, selectmode="browse")
        col_widths = [30, 60, 80, 270, 120]
        for c, w in zip(cols, col_widths):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="center" if w < 100 else "w")

        self.tree.tag_configure("High",   background="#FFC7CE")
        self.tree.tag_configure("Medium", background="#FFEB9C")
        self.tree.tag_configure("Low",    background="#C6EFCE")
        self.tree.tag_configure("Pass",   background="#E8F5E9")

        tree_sb = ttk.Scrollbar(card, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_sb.set)
        self.tree.grid(row=13, column=0, columnspan=2, sticky="nsew")
        tree_sb.grid(row=13, column=2, sticky="ns")

        card.columnconfigure(0, weight=1)
        card.columnconfigure(1, weight=1)
        card.rowconfigure(13, weight=1)

        tk.Label(self,
                 text="© Aruba OS Audit Tool  |  Based on CIS HPE Aruba Networking CX Switch Benchmark v1.0.1",
                 bg="#1F3864", fg="#6B8BB0", font=("Arial", 8)).pack(pady=6)

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select Aruba configuration log files",
            filetypes=[("Log/Config files", "*.log *.txt *.conf"), ("All files", "*.*")]
        )
        for p in paths:
            if p not in self._files:
                self._files.append(p)
                self.file_list.insert(tk.END, os.path.basename(p))
        self._update_file_count()

    def _remove_selected(self):
        selected = list(self.file_list.curselection())
        for idx in reversed(selected):
            self.file_list.delete(idx)
            self._files.pop(idx)
        self._update_file_count()

    def _clear_files(self):
        self.file_list.delete(0, tk.END)
        self._files.clear()
        self._update_file_count()

    def _update_file_count(self):
        n = len(self._files)
        self.file_count_var.set(f"{n} file(s) selected")

    def _browse_output(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.out_var.set(path)

    def _start_audit(self):
        if not self._files:
            messagebox.showwarning("No Files", "Please add at least one configuration log file.")
            return
        missing = [f for f in self._files if not os.path.isfile(f)]
        if missing:
            messagebox.showerror("File Not Found", "Cannot find:\n" + "\n".join(missing))
            return

        report_name = self.report_name_var.get().strip() or "AuditReport_MultiDevice"
        report_name = re.sub(r'[<>:"/\\\\|?*]', "_", report_name)
        if not report_name.endswith(".xlsx"):
            report_name += ".xlsx"

        out_dir = self.out_var.get()
        out_path = os.path.join(out_dir, report_name)

        self.run_btn.config(state="disabled", bg="#888")
        self.progress.start(12)
        self._clear_tree()
        self.status_var.set(f"Scanning {len(self._files)} file(s) against 61 CIS checks…")
        threading.Thread(target=self._audit_thread,
                         args=(list(self._files), out_path), daemon=True).start()

    def _audit_thread(self, filepaths, out_path):
        try:
            results = []
            all_findings = []
            for fp in filepaths:
                findings, device_ip = run_audit(fp)
                results.append((device_ip, findings))
                all_findings.extend(findings)
            write_multi_excel(results, out_path)
            self.after(0, self._audit_done, results, all_findings, out_path)
        except Exception:
            import traceback
            self.after(0, self._audit_error, traceback.format_exc())

    def _audit_done(self, results, all_findings, out_path):
        self.progress.stop()
        self.run_btn.config(state="normal", bg="#107C10")

        row_num = 1
        for device_ip, findings in results:
            if not findings:
                self.tree.insert("", "end",
                                 values=(row_num, "—", "✅ PASS", "All checks passed", device_ip),
                                 tags=("Pass",))
                row_num += 1
            else:
                for f in findings:
                    self.tree.insert("", "end",
                                     values=(row_num, f["cis_id"], f["severity"],
                                             f["observation"], f["device"]),
                                     tags=(f["severity"],))
                    row_num += 1

        total_issues = len(all_findings)
        devices_count = len(results)
        self.status_var.set(
            f"✅  Done. {devices_count} device(s) scanned, "
            f"{total_issues} finding(s) total.  "
            f"Report → {os.path.basename(out_path)}"
        )
        messagebox.showinfo(
            "Audit Complete",
            f"Scanned {devices_count} device(s) against 61 CIS checks\n"
            f"Total findings: {total_issues}\n\n"
            f"Report saved to:\n{out_path}"
        )

    def _audit_error(self, tb):
        self.progress.stop()
        self.run_btn.config(state="normal", bg="#107C10")
        self.status_var.set("❌  An error occurred. See details.")
        messagebox.showerror("Audit Error", f"An unexpected error occurred:\n\n{tb}")

    def _clear_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)


if __name__ == "__main__":
    app = AuditApp()
    app.mainloop()
