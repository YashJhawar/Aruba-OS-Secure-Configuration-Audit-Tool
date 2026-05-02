import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import os
import re
import threading
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

# ─────────────────────────────────────────────
# AUDIT KNOWLEDGE BASE
# ─────────────────────────────────────────────
AUDIT_CHECKS = [
    {
        "id": 1,
        "obs_title":  "Password complexity not configured appropriately",
        "observation": "It was observed that password complexity was not configured appropriately.",
        "severity": "High",
        "description": (
            "Enforces the Enterprise Password Policy by setting compliant local password "
            "requirements for the security appliance. The password policy helps to prevent "
            "unauthorized accesses by enforcing the password for more complexity and making "
            "them difficult to be guessed."
        ),
        "impact": (
            "Attackers can use techniques such as brute force, dictionary, hybrid etc to "
            "discover common or weak passwords that can lead to unauthorised access and may "
            "compromise the device."
        ),
        "recommendation": (
            "It is recommended to configure password complexity.\n\n"
            "Set a password complexity policy with a minimum of <X> characters long, includes at "
            "least <Y> lowercase/uppercase/special/numeric character, and doesn't allow users to "
            "reuse the past <Z> passwords:\n\n"
            "hostname(config)# password complexity\n"
            "hostname(config)# minimum-length <X>\n"
            "hostname(config)# lowercase-count <Y>\n"
            "hostname(config)# uppercase-count <Y>\n"
            "hostname(config)# special-char-count <Y>\n"
            "hostname(config)# numeric-count <Y>\n"
            "hostname(config)# history-count <Z>\n"
            "hostname(config)# enable"
        ),
        "expected_output": (
            "hostname(config)# password complexity\n"
            "hostname(config)# minimum-length <X>\n"
            "hostname(config)# lowercase-count <Y>\n"
            "hostname(config)# uppercase-count <Y>\n"
            "hostname(config)# special-char-count <Y>\n"
            "hostname(config)# numeric-count <Y>\n"
            "hostname(config)# history-count <Z>\n"
            "hostname(config)# enable"
        ),
        "check_fn": "check_password_complexity",
    },
    {
        "id": 2,
        "obs_title":  "Failed Attempts for Authentication Profiles not set",
        "observation": "It was observed that failed attempts is not configured for authentication profiles.",
        "severity": "Medium",
        "description": (
            "The 'Limiting Login Attempts' feature on Aruba AOS-CX switches is designed to "
            "enhance the security of the system by restricting the number of consecutive failed "
            "login attempts allowed for user accounts. Once the maximum number of failed attempts "
            "is reached, the account is blocked for a predetermined period - lockout-time."
        ),
        "impact": (
            "Absence of configured failed-attempt thresholds allows unlimited or excessive "
            "authentication attempts against administrative accounts. This significantly increases "
            "exposure to brute-force and password-guessing attacks, enabling an attacker to "
            "repeatedly attempt logins without restriction."
        ),
        "recommendation": (
            "It is recommended to enable login attempt failure limiting with a 60 second lockout.\n\n"
            "For the console interface (channel) only:\n"
            "hostname(config)# aaa authentication console-login-attempts 3 console-lockout-time 60\n\n"
            "For SSH/Telnet/https-server:\n"
            "hostname(config)# aaa authentication limit-login-attempts 3 lockout-time 60"
        ),
        "expected_output": (
            "hostname(config)# aaa authentication console-login-attempts 3 console-lockout-time 60\n\n"
            "For SSH/Telnet/https-server:\n"
            "hostname(config)# aaa authentication limit-login-attempts 3 lockout-time 60"
        ),
        "check_fn": "check_failed_attempts",
    },
    {
        "id": 3,
        "obs_title":  "NTP authentication not configured",
        "observation": "It was observed that NTP 'authentication' is not configured.",
        "severity": "Medium",
        "description": (
            "Enables NTP authentication in order to receive time information only from trusted "
            "sources. When authentication is not enabled, attackers can disguise as NTP servers "
            "and broadcast wrong time, and it will be difficult to correlate events upon an incident. "
            "In some other cases, attackers can perform NTP DDoS attacks such as NTP Amplification."
        ),
        "impact": (
            "When authentication is not enabled, attackers can disguise as NTP servers and broadcast "
            "wrong time, and it will be difficult to correlate events upon an incident. An attacker "
            "could potentially spoof NTP packets, leading to time-related attacks such as replay "
            "attacks or man-in-the-middle attacks."
        ),
        "recommendation": (
            "It is recommended to enable NTP authentication and set key to authentication NTP servers:\n\n"
            "hostname(config)# ntp authentication\n"
            "hostname(config)# ntp authentication-key 1 md5 <ntpauthkey>"
        ),
        "expected_output": (
            "hostname(config)# ntp authentication\n"
            "hostname(config)# ntp authentication-key 1 md5 <ntpauthkey>"
        ),
        "check_fn": "check_ntp_auth",
    },
    {
        "id": 4,
        "obs_title":  "SSH Allow List disabled",
        "observation": "It was observed that SSH allow list is disabled.",
        "severity": "Medium",
        "description": (
            "SSH server allow-list allows the switch admin to limit in-bound SSH connections to be "
            "from specific ip addresses or networks. Configure a list of addresses that will be the "
            "only hosts allowed to connect to the SSH servers running on all VRFs of the switch. "
            "By default, the allow-list is disabled and any host is allowed to connect given the "
            "correct authentication criteria."
        ),
        "impact": (
            "Disabling the SSH allow-list permits SSH access attempts from any IP address, increasing "
            "the attack surface for brute-force, credential guessing, or unauthorized remote access. "
            "Without IP-based restrictions, attackers on the same network segment or any reachable "
            "segment can target the management plane directly."
        ),
        "recommendation": (
            "It is recommended to ensure SSH Allow List is enabled.\n\n"
            "hostname(config)# ssh server allow-list\n"
            "hostname(config)# ip 1.1.1.1\n"
            "hostname(config)# enable"
        ),
        "expected_output": (
            "hostname(config)# ssh server allow-list\n"
            "hostname(config)# ip 1.1.1.1\n"
            "hostname(config)# enable"
        ),
        "check_fn": "check_ssh_allowlist",
    },
    {
        "id": 5,
        "obs_title":  "Syslog logging not configured",
        "observation": "It was observed that logging to a remote syslog server is not configured.",
        "severity": "Low",
        "description": (
            "Logging to a remote syslog server offers several advantages, including centralized log "
            "management, improved security, long-term data retention, and enhanced troubleshooting "
            "capabilities. By consolidating logs from various devices into a single location, "
            "administrators can easily monitor, analyze, and respond to issues across the network."
        ),
        "impact": (
            "Without a configured remote Syslog server, security-relevant events including "
            "authentication attempts, configuration changes, privilege escalations, and network "
            "anomalies are stored only locally on the switch. Local logs can be easily overwritten "
            "due to limited buffer size or lost entirely if the device is rebooted, cleared, or compromised."
        ),
        "recommendation": (
            "It is recommended to ensure syslog is configured.\n\n"
            "Configure the switch to send logs to a syslog-server using tls on the mgmt VRF and "
            "include auditable events:\n\n"
            "hostname(config)# logging <SYSLOG-SERVER> tls <PORT> auth-mode subject-name vrf mgmt "
            "include-auditable-events"
        ),
        "expected_output": (
            "hostname(config)# logging <SYSLOG-SERVER> tls <PORT> auth-mode subject-name vrf mgmt "
            "include-auditable-events"
        ),
        "check_fn": "check_syslog",
    },
]


# ─────────────────────────────────────────────
# AUDIT CHECK FUNCTIONS
# Each returns: (passed: bool, found_lines: list of (line_no, line_text))
# ─────────────────────────────────────────────

def _intent_scan(lines, intent_patterns):
    """
    Scan every line in the file against a list of intent patterns (case-insensitive).
    Returns sorted list of (line_no, line_text) for every line that matches ANY pattern.
    No context window — only lines that genuinely relate to the intent are returned.
    """
    matched = {}
    for i, line in enumerate(lines, 1):
        for pat in intent_patterns:
            if re.search(pat, line, re.IGNORECASE):
                matched[i] = line.rstrip()
                break   # one match per line is enough
    return sorted(matched.items())


# ── Intent patterns vs Pass patterns ─────────────────────────────────────────
#
# INTENT patterns:
#   Derived directly from the leading command keywords in each
#   "hostname(config)# <command>" line of the Recommendation.
#   They match any line in the log that relates to those commands —
#   whether complete, partial, misconfigured, or just commented out.
#   Rule: if the line contains the command keyword, it is evidence.
#
# PASS patterns:
#   Structurally complete version of the required command.
#   Used ONLY for pass/fail decision. Never for evidence collection.

# ── Check 1 – Password Complexity ────────────────────────────────────────────
# Recommendation commands: password complexity | minimum-length | lowercase-count
#   uppercase-count | special-char-count | numeric-count | history-count | enable
# Intent: any line containing these command keywords
_PWD_INTENT = [
    r"\bpassword\s+complexity\b",
    r"\bminimum-length\b",
    r"\blowercase-count\b",
    r"\buppercase-count\b",
    r"\bspecial-char-count\b",
    r"\bnumeric-count\b",
    r"\bhistory-count\b",
]
# Pass: ALL seven sub-commands must exist as actual config lines
_PWD_PASS = [
    r"^\s*password\s+complexity",
    r"^\s*minimum-length\s+\d+",
    r"^\s*lowercase-count\s+\d+",
    r"^\s*uppercase-count\s+\d+",
    r"^\s*special-char-count\s+\d+",
    r"^\s*numeric-count\s+\d+",
    r"^\s*history-count\s+\d+",
]

# ── Check 2 – Failed Login Attempts ──────────────────────────────────────────
# Recommendation commands:
#   aaa authentication console-login-attempts <N> console-lockout-time <T>
#   aaa authentication limit-login-attempts <N> lockout-time <T>
# Intent: any line containing these command keywords
_AAA_INTENT = [
    r"\baaa\s+authentication\s+console-login-attempts\b",
    r"\baaa\s+authentication\s+limit-login-attempts\b",
    r"\bconsole-login-attempts\b",
    r"\blimit-login-attempts\b",
]
# Pass: BOTH commands must be fully and correctly configured
_AAA_PASS_CONSOLE = r"aaa\s+authentication\s+console-login-attempts\s+\d+\s+console-lockout-time\s+\d+"
_AAA_PASS_LIMIT   = r"aaa\s+authentication\s+limit-login-attempts\s+\d+\s+lockout-time\s+\d+"

# ── Check 3 – NTP Authentication ─────────────────────────────────────────────
# Recommendation commands:
#   ntp authentication
#   ntp authentication-key 1 md5 <key>
# Intent: any line containing these command keywords
_NTP_INTENT = [
    r"\bntp\s+authentication\b",
    r"\bntp\s+authentication-key\b",
]
# Pass: BOTH lines must be present
_NTP_PASS_AUTH = r"^\s*ntp\s+authentication\s*$"
_NTP_PASS_KEY  = r"^\s*ntp\s+authentication-key\s+"

# ── Check 4 – SSH Allow List ──────────────────────────────────────────────────
# Recommendation commands:
#   ssh server allow-list
#   ip <address>
#   enable
# Intent: any line containing these command keywords
_SSH_INTENT = [
    r"\bssh\s+server\s+allow-list\b",
    r"\bssh\s+server\b",
]
# Pass: block + at least one IP + enable must all be present in sequence
_SSH_PASS_BLOCK  = r"ssh\s+server\s+allow-list"
_SSH_PASS_IP     = r"^\s*ip\s+\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}"
_SSH_PASS_ENABLE = r"^\s*enable\s*$"

# ── Check 5 – Syslog ─────────────────────────────────────────────────────────
# Recommendation command (exact structure required):
#   logging <SYSLOG-SERVER> tls <PORT> auth-mode subject-name vrf mgmt include-auditable-events
# Intent: only lines that contain ALL key structural tokens of this command.
#   A plain "logging <ip> udp 514" does NOT match — must have tls + auth-mode + vrf mgmt + include-auditable-events.
#   If none found → Missing Configuration.
_SYSLOG_INTENT = [
    r"\blogging\s+\S+\s+tls\s+\S+\s+auth-mode\s+subject-name\s+vrf\s+mgmt\s+include-auditable-events\b",
]
# Pass: same exact structural pattern
_SYSLOG_PASS = r"^\s*logging\s+\S+\s+tls\s+\S+\s+auth-mode\s+subject-name\s+vrf\s+mgmt\s+include-auditable-events"


# ── Check functions ───────────────────────────────────────────────────────────

def check_password_complexity(lines):
    hits = {pat: False for pat in _PWD_PASS}
    for line in lines:
        for pat in _PWD_PASS:
            if not hits[pat] and re.search(pat, line, re.IGNORECASE):
                hits[pat] = True
    if all(hits.values()):
        return True, []
    return False, _intent_scan(lines, _PWD_INTENT)


def check_failed_attempts(lines):
    has_console = any(re.search(_AAA_PASS_CONSOLE, l, re.IGNORECASE) for l in lines)
    has_limit   = any(re.search(_AAA_PASS_LIMIT,   l, re.IGNORECASE) for l in lines)
    if has_console and has_limit:
        return True, []
    return False, _intent_scan(lines, _AAA_INTENT)


def check_ntp_auth(lines):
    has_auth = any(re.search(_NTP_PASS_AUTH, l, re.IGNORECASE) for l in lines)
    has_key  = any(re.search(_NTP_PASS_KEY,  l, re.IGNORECASE) for l in lines)
    if has_auth and has_key:
        return True, []
    return False, _intent_scan(lines, _NTP_INTENT)


def check_ssh_allowlist(lines):
    has_block  = any(re.search(_SSH_PASS_BLOCK,  l, re.IGNORECASE) for l in lines)
    has_ip     = any(re.search(_SSH_PASS_IP,     l, re.IGNORECASE) for l in lines)
    has_enable = any(re.search(_SSH_PASS_ENABLE, l, re.IGNORECASE) for l in lines)
    if has_block and has_ip and has_enable:
        return True, []
    return False, _intent_scan(lines, _SSH_INTENT)


def check_syslog(lines):
    if any(re.search(_SYSLOG_PASS, l, re.IGNORECASE) for l in lines):
        return True, []
    return False, _intent_scan(lines, _SYSLOG_INTENT)


CHECK_DISPATCH = {
    "check_password_complexity": check_password_complexity,
    "check_failed_attempts":     check_failed_attempts,
    "check_ntp_auth":            check_ntp_auth,
    "check_ssh_allowlist":       check_ssh_allowlist,
    "check_syslog":              check_syslog,
}


# ─────────────────────────────────────────────
# HELPER: extract device IP from filename
# ─────────────────────────────────────────────

def extract_device_ip(filepath, lines):
    # Try filename first
    basename = os.path.basename(filepath)
    ip_match = re.search(r"(\d{1,3}(?:\.\d{1,3}){3})", basename)
    if ip_match:
        return ip_match.group(1)
    # Try first 10 lines of file
    for line in lines[:10]:
        ip_match = re.search(r"(\d{1,3}(?:\.\d{1,3}){3})", line)
        if ip_match:
            return ip_match.group(1)
    return "Unknown"


# ─────────────────────────────────────────────
# CORE AUDIT RUNNER
# ─────────────────────────────────────────────

def run_audit(filepath):
    with open(filepath, "r", errors="replace") as f:
        lines = f.readlines()

    device_ip = extract_device_ip(filepath, lines)
    findings  = []

    for check in AUDIT_CHECKS:
        fn = CHECK_DISPATCH[check["check_fn"]]
        passed, found_lines = fn(lines)

        if passed:
            continue  # No finding to report

        # ── Build Received Output ──────────────────────────────────────────
        # Rules:
        #   1. First line is always: "The config line that needs to be checked:"
        #   2. If matching lines exist → Format 1:
        #        hostname(config)# <actual line from log file>
        #        (one per matched line, raw text preserved)
        #        Reference Line: <first line no> - <last line no>  (or single no)
        #   3. If no matching lines at all → Format 2:
        #        Missing Configuration
        #
        # Filtering: exclude lines that are ONLY whitespace, ONLY "!",
        # or decorative section-header comments (e.g. "! ── Section ──────").
        # Keep all genuine config lines AND meaningful comment lines that
        # contain actual command keywords (they are evidence).

        def _is_meaningful(raw):
            s = raw.strip()
            if not s:
                return False                          # blank line
            if s == "!":
                return False                          # lone separator
            # Decorative section headers: "! ── text ──" or "! === text ==="
            if re.match(r"^!\s*[─═\-]{2,}", s):
                return False
            return True

        def _display_text(raw):
            # Return the line exactly as it appears in the file (stripped of
            # leading/trailing whitespace only). Do NOT strip "!" comment chars —
            # the full raw line is the evidence.
            return raw.strip()

        meaningful = [
            (ln, lt) for ln, lt in found_lines
            if _is_meaningful(lt)
        ]

        if not meaningful:
            received = (
                "The config line that needs to be checked:\n"
                "\n"
                "Missing Configuration"
            )
        else:
            config_lines = "\n".join(
                f"hostname(config)# {_display_text(lt)}" for _, lt in meaningful
            )
            line_nums = [ln for ln, _ in meaningful]
            ref = (
                f"Reference Line: {line_nums[0]}"
                if len(line_nums) == 1
                else f"Reference Line: {line_nums[0]} - {line_nums[-1]}"
            )
            received = (
                "The config line that needs to be checked:\n"
                "\n"
                f"{config_lines}\n"
                "\n"
                f"{ref}"
            )

        # ── Build Expected Output ──────────────────────────────────────────
        # First line is always: "The expected config line that needs to be altered:"
        # followed by 2 blank lines, then the expected config content.
        expected = (
            "The expected config line that needs to be altered:\n"
            "\n"
            "\n"
            f"{check['expected_output']}"
        )

        findings.append({
            "severity":    check["severity"],
            "obs_title":   check["obs_title"],
            "observation": check["observation"],
            "device":      device_ip,
            "description": check["description"],
            "impact":      check["impact"],
            "recommendation": check["recommendation"],
            "received":    received,
            "expected":    expected,
        })

    return findings, device_ip


# ─────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────

# ── Formatting constants ──────────────────────────────────────────────────────

# Rule 1 – Row 1: bold text, yellow background
HEADER_FILL = PatternFill("solid", fgColor="FFFF00")          # yellow
HEADER_FONT = Font(bold=True, color="000000", name="Arial", size=10)

# Rule 1 – Data rows: no bold
CELL_FONT = Font(bold=False, name="Arial", size=9)

# Rule 4 – Row 1 centre-aligned; data rows top-left aligned
HEADER_ALIGN = Alignment(horizontal="center", vertical="center",
                         wrap_text=True)
DATA_ALIGN   = Alignment(horizontal="left", vertical="top",
                         wrap_text=True)

# Rule 7 – All borders on every cell
_THIN   = Side(style="thin", color="000000")
BORDER  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Rule 5 – No fill for data rows (white / transparent)
NO_FILL = PatternFill(fill_type=None)

# Rule 6 – Column B severity colours (only column B gets colour)
SEV_FILL = {
    "High":   PatternFill("solid", fgColor="FF0000"),   # red
    "Medium": PatternFill("solid", fgColor="F5C242"),   # hex F5C242
    "Low":    PatternFill("solid", fgColor="00B050"),   # green #00B050
}

# Columns A (#), B (Severity), D (Affected Device) → centre-top aligned
# All other data columns → top-left aligned
CENTER_TOP_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)

COLUMNS = [
    ("#",               6),
    ("Severity",       12),
    ("Observation",    30),
    ("Affected Device",18),
    ("Description",    40),
    ("Impact",         40),
    ("Recommendation", 45),
    ("Received Output",45),
    ("Expected Output",45),
    ("FSL Remarks",    20),
    ("EY Remarks",     20),
]


def _write_sheet(wb, findings, sheet_name):
    """Write one device's findings into a new sheet in wb."""

    # ── Safe sheet name (Excel: max 31 chars, no special chars) ──────────────
    safe_name = re.sub(r'[\\/*?:\[\]]', '_', sheet_name)[:31]
    existing  = [s.title for s in wb.worksheets]
    if safe_name in existing:
        suffix = 2
        while f"{safe_name[:28]}_{suffix}" in existing:
            suffix += 1
        safe_name = f"{safe_name[:28]}_{suffix}"

    ws = wb.create_sheet(title=safe_name)

    # Rule 2 – Hide gridlines
    ws.sheet_view.showGridLines = False

    # ── Row 1 – Header ────────────────────────────────────────────────────────
    for col_idx, (header, width) in enumerate(COLUMNS, 1):
        cell            = ws.cell(row=1, column=col_idx, value=header)
        cell.font       = HEADER_FONT       # Rule 1 – bold
        cell.fill       = HEADER_FILL       # Rule 1 – yellow background
        cell.alignment  = HEADER_ALIGN      # Rule 4 – centre aligned
        cell.border     = BORDER            # Rule 7 – all borders
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30        # Rule 3 – row 1 height = 30
    ws.freeze_panes = "A2"

    # ── No-findings notice ────────────────────────────────────────────────────
    if not findings:
        ws.cell(row=2, column=1,
                value="All checks PASSED — no issues found for this device.")
        ws.cell(row=2, column=1).font = Font(name="Arial", size=10,
                                             bold=True, color="107C10")
        ws.row_dimensions[2].height = 30
        return

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, f in enumerate(findings, 2):

        # Build rich text for Observation cell (col C):
        # Line 1: bold title
        # Two blank lines
        # Line 4: normal weight detail sentence
        obs_rich = CellRichText(
            TextBlock(InlineFont(b=True,  rFont="Arial", sz=9), f["obs_title"]),
            TextBlock(InlineFont(b=False, rFont="Arial", sz=9), "\n\n\n" + f["observation"]),
        )

        values = [
            row_idx - 1,            # col A  #
            f["severity"],          # col B  Severity      ← bold applied separately
            obs_rich,               # col C  Observation   ← rich text
            f["device"],            # col D  Affected Device
            f["description"],       # col E  Description
            f["impact"],            # col F  Impact
            f["recommendation"],    # col G  Recommendation
            f["received"],          # col H  Received Output
            f["expected"],          # col I  Expected Output
            "",                     # col J  FSL Remarks
            "",                     # col K  EY Remarks
        ]

        for col_idx, val in enumerate(values, 1):
            cell        = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = BORDER         # Rule 7 – all borders
            cell.fill   = NO_FILL        # Rule 5 – no background

            # Column B: bold severity text + severity fill colour
            if col_idx == 2:
                cell.font      = Font(bold=True, name="Arial", size=9)
                cell.fill      = SEV_FILL.get(f["severity"], NO_FILL)
                cell.alignment = CENTER_TOP_ALIGN
            # Column C: rich text already set — only set alignment, no font override
            elif col_idx == 3:
                cell.alignment = DATA_ALIGN
            # Columns A and D: centre-top, normal font
            elif col_idx in (1, 4):
                cell.font      = CELL_FONT
                cell.alignment = CENTER_TOP_ALIGN
            # All other data columns: top-left, normal font
            else:
                cell.font      = CELL_FONT
                cell.alignment = DATA_ALIGN

        ws.row_dimensions[row_idx].height = 180   # Rule 3 – data rows = 180


def write_multi_excel(results, out_path):
    """
    results: list of (device_ip, findings) tuples — one per scanned file.
    One sheet per device, tab named by IP address.
    """
    wb = Workbook()
    wb.remove(wb.active)          # drop the default empty sheet

    for device_ip, findings in results:
        _write_sheet(wb, findings, device_ip)

    wb.save(out_path)


def write_excel(findings, out_path, device_ip):
    """Single-file convenience wrapper."""
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, findings, device_ip)
    wb.save(out_path)


# ─────────────────────────────────────────────
# TKINTER GUI
# ─────────────────────────────────────────────

class AuditApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Aruba OS – Secure Configuration Auditor")
        self.resizable(True, True)
        self.minsize(620, 620)
        self.configure(bg="#1F3864")
        self._files = []          # list of selected file paths
        self._build_ui()

    # ── UI Construction ──────────────────────

    def _build_ui(self):
        # ─ Title bar ─
        title_frame = tk.Frame(self, bg="#1F3864", pady=14)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame,
            text="🔒  Aruba OS  |  Secure Configuration Auditor",
            bg="#1F3864", fg="white",
            font=("Arial", 16, "bold"),
        ).pack()
        tk.Label(
            title_frame,
            text="Automated security checklist review for AOS-CX devices",
            bg="#1F3864", fg="#A0B4D0",
            font=("Arial", 9),
        ).pack()

        # ─ Card container ─
        card = tk.Frame(self, bg="white", padx=28, pady=24, relief="flat")
        card.pack(padx=24, pady=(0, 24), fill="both", expand=True)

        # ── File Selection ──────────────────────────────────────────────────
        tk.Label(card, text="Configuration Log Files", bg="white",
                 font=("Arial", 10, "bold"), anchor="w").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 4))

        # Listbox showing selected files
        list_frame = tk.Frame(card, bg="white")
        list_frame.grid(row=1, column=0, columnspan=3, sticky="nsew", pady=(0, 4))

        self.file_list = tk.Listbox(
            list_frame, height=6, font=("Arial", 9),
            relief="solid", bd=1, bg="#F4F6F8",
            selectmode=tk.EXTENDED, activestyle="none"
        )
        list_sb = ttk.Scrollbar(list_frame, orient="vertical",
                                command=self.file_list.yview)
        self.file_list.configure(yscrollcommand=list_sb.set)
        self.file_list.pack(side="left", fill="both", expand=True)
        list_sb.pack(side="right", fill="y")

        # File action buttons
        btn_frame = tk.Frame(card, bg="white")
        btn_frame.grid(row=2, column=0, columnspan=3, sticky="w", pady=(0, 6))

        tk.Button(
            btn_frame, text="＋  Add Files", command=self._add_files,
            bg="#2E5BA8", fg="white", font=("Arial", 9, "bold"),
            relief="flat", cursor="hand2", padx=10, pady=4
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            btn_frame, text="✕  Remove Selected", command=self._remove_selected,
            bg="#C0392B", fg="white", font=("Arial", 9, "bold"),
            relief="flat", cursor="hand2", padx=10, pady=4
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            btn_frame, text="⊘  Clear All", command=self._clear_files,
            bg="#7F8C8D", fg="white", font=("Arial", 9, "bold"),
            relief="flat", cursor="hand2", padx=10, pady=4
        ).pack(side="left")

        self.file_count_var = tk.StringVar(value="0 file(s) selected")
        tk.Label(card, textvariable=self.file_count_var, bg="white",
                 fg="#888", font=("Arial", 8, "italic")).grid(
            row=3, column=0, columnspan=3, sticky="w")

        # ── Output Folder ───────────────────────────────────────────────────
        tk.Label(card, text="Output Folder", bg="white",
                 font=("Arial", 10, "bold"), anchor="w").grid(
            row=4, column=0, columnspan=3, sticky="w", pady=(14, 4))

        self.out_var = tk.StringVar(
            value=os.path.expanduser("~\\Desktop")
            if os.name == "nt" else os.path.expanduser("~/Desktop")
        )
        tk.Entry(card, textvariable=self.out_var, width=48,
                 font=("Arial", 9), relief="solid", bd=1).grid(
            row=5, column=0, columnspan=2, sticky="ew", ipady=4)

        tk.Button(
            card, text="  Browse…  ", command=self._browse_output,
            bg="#2E5BA8", fg="white", font=("Arial", 9, "bold"),
            relief="flat", cursor="hand2", padx=8, pady=4
        ).grid(row=5, column=2, padx=(8, 0))

        # ── Report Name ─────────────────────────────────────────────────────
        tk.Label(card, text="Output Report Name", bg="white",
                 font=("Arial", 10, "bold"), anchor="w").grid(
            row=6, column=0, columnspan=3, sticky="w", pady=(12, 4))

        self.report_name_var = tk.StringVar(value="AuditReport_MultiDevice")
        tk.Entry(card, textvariable=self.report_name_var, width=48,
                 font=("Arial", 9), relief="solid", bd=1).grid(
            row=7, column=0, columnspan=2, sticky="ew", ipady=4)
        tk.Label(card, text=".xlsx", bg="white",
                 font=("Arial", 9), fg="#555").grid(row=7, column=2, sticky="w", padx=(6,0))

        # ── Separator + Run ─────────────────────────────────────────────────
        ttk.Separator(card, orient="horizontal").grid(
            row=8, column=0, columnspan=3, sticky="ew", pady=18)

        self.run_btn = tk.Button(
            card, text="▶   Run Audit", command=self._start_audit,
            bg="#107C10", fg="white", font=("Arial", 11, "bold"),
            relief="flat", cursor="hand2", padx=16, pady=8, width=20
        )
        self.run_btn.grid(row=9, column=0, columnspan=3)

        # ── Progress / Status ───────────────────────────────────────────────
        self.progress = ttk.Progressbar(card, mode="indeterminate", length=460)
        self.progress.grid(row=10, column=0, columnspan=3, pady=(14, 4))

        self.status_var = tk.StringVar(value="Ready. Add .log files to begin.")
        tk.Label(card, textvariable=self.status_var, bg="white",
                 fg="#555", font=("Arial", 9, "italic")).grid(
            row=11, column=0, columnspan=3, pady=(0, 6))

        # ── Findings Preview Table ──────────────────────────────────────────
        tk.Label(card, text="Findings Preview", bg="white",
                 font=("Arial", 10, "bold"), anchor="w").grid(
            row=12, column=0, columnspan=3, sticky="w", pady=(10, 4))

        cols = ("#", "Severity", "Observation", "Affected Device")
        self.tree = ttk.Treeview(card, columns=cols, show="headings",
                                 height=7, selectmode="browse")
        col_widths = [30, 80, 280, 130]
        for c, w in zip(cols, col_widths):
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, anchor="center" if w < 100 else "w")

        self.tree.tag_configure("High",   background="#FFC7CE")
        self.tree.tag_configure("Medium", background="#FFEB9C")
        self.tree.tag_configure("Low",    background="#C6EFCE")
        self.tree.tag_configure("Pass",   background="#E8F5E9")

        tree_sb = ttk.Scrollbar(card, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_sb.set)
        self.tree.grid(row=13, column=0, columnspan=2, sticky="nsew")
        tree_sb.grid(row=13, column=2, sticky="ns")

        card.columnconfigure(0, weight=1)
        card.columnconfigure(1, weight=1)
        card.rowconfigure(13, weight=1)

        # Footer
        tk.Label(self, text="© Aruba OS Audit Tool  |  For internal security review use only",
                 bg="#1F3864", fg="#6B8BB0", font=("Arial", 8)).pack(pady=6)

    # ── File Management ───────────────────────

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select Aruba configuration log files",
            filetypes=[("Log/Config files", "*.log *.txt *.conf"), ("All files", "*.*")]
        )
        for p in paths:
            if p not in self._files:
                self._files.append(p)
                self.file_list.insert(tk.END, os.path.basename(p))
        self._update_file_count()

    def _remove_selected(self):
        selected = list(self.file_list.curselection())
        for idx in reversed(selected):
            self.file_list.delete(idx)
            self._files.pop(idx)
        self._update_file_count()

    def _clear_files(self):
        self.file_list.delete(0, tk.END)
        self._files.clear()
        self._update_file_count()

    def _update_file_count(self):
        n = len(self._files)
        self.file_count_var.set(f"{n} file(s) selected")

    def _browse_output(self):
        path = filedialog.askdirectory(title="Select output folder")
        if path:
            self.out_var.set(path)

    # ── Audit Execution ───────────────────────

    def _start_audit(self):
        if not self._files:
            messagebox.showwarning("No Files", "Please add at least one configuration log file.")
            return
        missing = [f for f in self._files if not os.path.isfile(f)]
        if missing:
            messagebox.showerror("File Not Found",
                                 "Cannot find:\n" + "\n".join(missing))
            return

        report_name = self.report_name_var.get().strip() or "AuditReport_MultiDevice"
        # Sanitise report name
        report_name = re.sub(r'[<>:"/\\\\|?*]', "_", report_name)
        if not report_name.endswith(".xlsx"):
            report_name += ".xlsx"

        out_dir  = self.out_var.get()
        out_path = os.path.join(out_dir, report_name)

        self.run_btn.config(state="disabled", bg="#888")
        self.progress.start(12)
        self._clear_tree()
        self.status_var.set(f"Scanning {len(self._files)} file(s)…")
        threading.Thread(target=self._audit_thread,
                         args=(list(self._files), out_path), daemon=True).start()

    def _audit_thread(self, filepaths, out_path):
        try:
            results      = []   # (device_ip, findings)
            all_findings = []

            for fp in filepaths:
                findings, device_ip = run_audit(fp)
                results.append((device_ip, findings))
                all_findings.extend(findings)

            write_multi_excel(results, out_path)
            self.after(0, self._audit_done, results, all_findings, out_path)

        except Exception:
            import traceback
            self.after(0, self._audit_error, traceback.format_exc())

    def _audit_done(self, results, all_findings, out_path):
        self.progress.stop()
        self.run_btn.config(state="normal", bg="#107C10")

        row_num = 1
        for device_ip, findings in results:
            if not findings:
                self.tree.insert("", "end",
                                 values=(row_num, "✅ PASS",
                                         "All checks passed", device_ip),
                                 tags=("Pass",))
                row_num += 1
            else:
                for f in findings:
                    self.tree.insert("", "end",
                                     values=(row_num, f["severity"],
                                             f["observation"], f["device"]),
                                     tags=(f["severity"],))
                    row_num += 1

        total_issues  = len(all_findings)
        devices_count = len(results)
        self.status_var.set(
            f"✅  Done. {devices_count} device(s) scanned, "
            f"{total_issues} finding(s) total.  "
            f"Report → {os.path.basename(out_path)}"
        )
        messagebox.showinfo(
            "Audit Complete",
            f"Scanned {devices_count} device(s)\n"
            f"Total findings: {total_issues}\n\n"
            f"Report saved to:\n{out_path}"
        )

    def _audit_error(self, tb):
        self.progress.stop()
        self.run_btn.config(state="normal", bg="#107C10")
        self.status_var.set("❌  An error occurred. See details.")
        messagebox.showerror("Audit Error",
                             f"An unexpected error occurred:\n\n{tb}")

    def _clear_tree(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
if __name__ == "__main__":
    app = AuditApp()
    app.mainloop()
